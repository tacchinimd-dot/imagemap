"""
make_fit_overrides.py — 7개 탭(아우터/폴로/티셔츠반팔/스커트/맨투맨/스웨터/팬츠쇼츠)에 속한
모든 의류 상품에 대해 수기 핏 입력 엑셀 파일을 생성.

출력: C:\\Users\\AD0903\\brand_crawler\\fit_overrides.xlsx

기존 파일이 있으면 '새 핏 ★' 열의 기존 입력값을 보존.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding="utf-8")

# line_matrix 로직 재사용
from line_matrix import (
    BRAND_LABEL, LINES, TABS,
    classify_fit, classify_line, classify_tab,
    load_alo, load_lacoste, load_rl, load_wilson,
    load_descente, load_lululemon, load_celine, load_loropiana,
    load_skims, load_miumiu, load_prada, load_sportyandrich,
)

ROOT = Path(r"C:\Users\AD0903\brand_crawler")       # 입력(raw 데이터) 위치
SCRIPT_DIR = Path(__file__).resolve().parent        # 출력(xlsx) 위치
OUT = SCRIPT_DIR / "fit_overrides.xlsx"

TAB_LABEL = {k: lbl for k, lbl, _ in TABS}
LINE_LABEL = {k: lbl for k, lbl, _, _ in LINES}
# 라인별 색상 톤(클래식=연녹/어슬레져=연파/스포츠=연빨)
BRAND_COLOR = {
    "alo": "DBEAFE", "lululemon": "BFDBFE",
    "wilson": "FEE2E2", "descente": "FECACA",
    "lacoste": "DCFCE7", "rl": "DCFCE7",
    "celine": "BBF7D0", "loropiana": "A7F3D0",
    "skims": "DBEAFE", "miumiu": "C7D2FE", "prada": "C7D2FE",
    "sportyandrich": "DCFCE7",
}

FIT_LABELS = {
    "slim": "슬림핏", "regular": "레귤러핏", "over": "세미오버핏~오버핏",
    "fitted": "피티드", "flare": "플레어", "pleats": "플리츠", "widepleats": "와이드플리츠",
    "straight": "스트레이트핏", "wide": "세미와이드~와이드핏",
    "mini": "미니", "midi": "미디", "maxi": "맥시",
}
DROPDOWN = ["슬림핏", "레귤러핏", "세미오버핏~오버핏"]
DROPDOWN_SKIRT = ["피티드", "플레어", "플리츠", "와이드플리츠"]
DROPDOWN_BOTTOM = ["슬림핏", "스트레이트핏", "세미와이드~와이드핏"]
DROPDOWN_DRESS = ["미니", "미디", "맥시"]


def load_existing_new_fits() -> dict[tuple[str, str], str]:
    """기존 파일이 있으면 '새 핏 ★' 입력값 보존 (key = (브랜드라벨, 상품URL))."""
    if not OUT.exists():
        return {}
    wb = load_workbook(str(OUT), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}
    h = [str(x).strip() if x else "" for x in rows[1]]  # row2 = 헤더
    try:
        bi = h.index("브랜드"); ui = h.index("상품 URL"); ni = h.index("새 핏 ★")
    except ValueError:
        return {}
    out = {}
    for r in rows[2:]:
        if not r or not r[bi]:
            continue
        v = str(r[ni] or "").strip()
        if v:
            out[(str(r[bi]).strip(), str(r[ui] or "").strip())] = v
    return out


def main():
    items = (
        load_alo() + load_wilson() + load_lacoste() + load_rl()
        + load_descente() + load_lululemon() + load_celine() + load_loropiana()
        + load_skims() + load_miumiu() + load_prada() + load_sportyandrich()
    )
    existing = load_existing_new_fits()

    # 7개 탭 대상 상품만 추출
    rows = []
    for it in items:
        tab = classify_tab(it)
        if not tab:
            continue
        line = classify_line(it["brand"])
        if not line:
            continue
        fit_key = classify_fit(it, overrides={}, tab=tab)
        cur_fit = FIT_LABELS.get(fit_key, "핏값없음")
        rows.append({
            "brand_key": it["brand"],
            "brand_label": BRAND_LABEL[it["brand"]],
            "line": LINE_LABEL[line],
            "tab_key": tab,
            "tab": TAB_LABEL[tab],
            "sub_raw": it["sub_raw"] or "(없음)",
            "name": it["name"],
            "cur_fit": cur_fit,
            "image_url": it["image_url"],
            "product_url": it["product_url"],
        })

    # 정렬: 탭 순서 → 라인 순서 → 브랜드 → 상품명
    tab_order = {lbl: i for i, (_, lbl, _) in enumerate(TABS)}
    line_order = {lbl: i for i, (_, lbl, _, _) in enumerate(LINES)}
    rows.sort(key=lambda r: (tab_order[r["tab"]], line_order[r["line"]], r["brand_label"], r["name"]))

    # 작성
    wb = Workbook()
    ws = wb.active
    ws.title = "fit_overrides"

    headers = ["브랜드", "라인", "탭", "소분류(원본)", "상품명",
               "현재 핏(분류결과)", "새 핏 ★", "이미지 URL", "상품 URL"]

    # 1행: 안내 (DO NOT MOVE — line_matrix.py가 row1을 안내로, row2를 헤더로 읽음)
    ws.cell(1, 1, "※ '새 핏 ★' 열에만 입력하세요. 빈 칸은 현재 핏(분류결과)을 그대로 사용합니다. "
                  "상의 탭: 슬림핏/레귤러핏/세미오버핏~오버핏 | 스커트 탭: 논플리츠/잔플리츠/와이드플리츠 | "
                  "팬츠/쇼츠 탭: 슬림핏/스트레이트핏/세미와이드~와이드핏")
    ws.cell(1, 1).font = Font(bold=True, color="B45309")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="FEF3C7")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # 2행: 헤더
    for i, h in enumerate(headers, 1):
        c = ws.cell(2, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F2937")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # 3행~: 데이터
    for ri, r in enumerate(rows, start=3):
        ws.cell(ri, 1, r["brand_label"]).fill = PatternFill("solid", fgColor=BRAND_COLOR[r["brand_key"]])
        ws.cell(ri, 2, r["line"])
        ws.cell(ri, 3, r["tab"])
        ws.cell(ri, 4, r["sub_raw"])
        ws.cell(ri, 5, r["name"])
        cur = ws.cell(ri, 6, r["cur_fit"])
        if r["cur_fit"] == "핏값없음":
            cur.fill = PatternFill("solid", fgColor="FEF3C7")
            cur.font = Font(bold=True, color="B45309")
        # 새 핏 ★ — 기존 입력값 보존
        prev = existing.get((r["brand_label"], r["product_url"]), "")
        nf = ws.cell(ri, 7, prev)
        nf.fill = PatternFill("solid", fgColor="FFFBEB")
        nf.alignment = Alignment(horizontal="center")
        ws.cell(ri, 8, r["image_url"])
        url_cell = ws.cell(ri, 9, r["product_url"])
        url_cell.hyperlink = r["product_url"]
        url_cell.font = Font(color="2563EB", underline="single")

    # 드롭다운: 새 핏 ★ 열 (G) — 탭별로 다른 드롭다운
    # 일반 탭: 슬림/레귤러/세미오버~오버
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(DROPDOWN) + '"',
        allow_blank=True,
    )
    dv.error = "드롭다운 값 중 하나를 선택하세요"
    dv.errorTitle = "잘못된 값"
    ws.add_data_validation(dv)

    # 스커트 탭: 논플리츠/잔플리츠/와이드플리츠
    dv_sk = DataValidation(
        type="list",
        formula1='"' + ",".join(DROPDOWN_SKIRT) + '"',
        allow_blank=True,
    )
    dv_sk.error = "드롭다운 값 중 하나를 선택하세요 (스커트)"
    dv_sk.errorTitle = "잘못된 값"
    ws.add_data_validation(dv_sk)

    # 팬츠/쇼츠 탭: 슬림핏/스트레이트핏/세미와이드~와이드핏
    dv_bt = DataValidation(
        type="list",
        formula1='"' + ",".join(DROPDOWN_BOTTOM) + '"',
        allow_blank=True,
    )
    dv_bt.error = "드롭다운 값 중 하나를 선택하세요 (팬츠/쇼츠)"
    dv_bt.errorTitle = "잘못된 값"
    ws.add_data_validation(dv_bt)

    # row별로 탭에 맞는 DV에 추가
    for ri, r in enumerate(rows, start=3):
        cell_ref = f"G{ri}"
        if r["tab_key"] == "skirt":
            dv_sk.add(cell_ref)
        elif r["tab_key"] == "bottom":
            dv_bt.add(cell_ref)
        else:
            dv.add(cell_ref)

    # 열 너비
    widths = [14, 10, 16, 18, 52, 18, 18, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 고정: 헤더행
    ws.freeze_panes = "A3"

    wb.save(str(OUT))

    total = len(rows)
    no_fit = sum(1 for r in rows if r["cur_fit"] == "핏값없음")
    print(f"[done] {OUT}")
    print(f"  총 {total}건 (현재 핏 분류됨 {total - no_fit}건 · 핏값없음 {no_fit}건)")
    print(f"  브랜드별: " + ", ".join(
        f"{BRAND_LABEL[b]}={sum(1 for r in rows if r['brand_key']==b)}"
        for b in ("alo", "lululemon", "skims", "miumiu", "prada",
                  "wilson", "descente",
                  "lacoste", "rl", "celine", "loropiana", "sportyandrich")
    ))
    print(f"  탭별:  " + ", ".join(
        f"{lbl}={sum(1 for r in rows if r['tab']==lbl)}" for _, lbl, _ in TABS
    ))


if __name__ == "__main__":
    main()
