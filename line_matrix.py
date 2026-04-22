"""
line_matrix.py — 라인(클래식/어슬레져/스포츠) × 핏(슬림/레귤러/세미오버~오버) 매트릭스 대시보드 생성.
탭 7개: 아우터 · 폴로 · 티셔츠(반팔) · 스커트 · 맨투맨 및 후디 · 스웨터 · 팬츠 및 쇼츠

출력:
    C:\\Users\\AD0903\\brand_crawler\\line_matrix.html
"""
from __future__ import annotations

import html
import json
import os
import sys
from collections import defaultdict
from glob import glob
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

# ── 로컬/원격 이미지 → data URL 변환 (캡처 CORS 회피) ──────────────────
import base64 as _b64
import io as _io
import hashlib as _hashlib
try:
    from PIL import Image as _PILImage
    _PIL_OK = True
except ImportError:
    _PIL_OK = False
try:
    import httpx as _httpx
    _HTTPX_OK = True
except ImportError:
    _HTTPX_OK = False

_DATA_URL_CACHE: dict[str, str] = {}
_IMAGE_CACHE_DIR = Path(r"C:\Users\AD0903\brand_crawler") / ".image_cache"

# 1×1 진짜 투명 PNG (color type 6, alpha=0)
_PLACEHOLDER = ("data:image/png;base64,"
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNkYAAAAAYAAjCB0C8AAAAASUVORK5CYII=")

_FAILED_URLS: list[str] = []


def _pil_to_data_url(img, max_w: int = 260, q: int = 78) -> str:
    if img.mode in ("RGBA", "LA"):
        bg = _PILImage.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[-1])
        img = bg
    else:
        img = img.convert("RGB")
    if img.width > max_w:
        ratio = max_w / img.width
        img = img.resize((max_w, int(img.height * ratio)), _PILImage.LANCZOS)
    buf = _io.BytesIO()
    img.save(buf, format="JPEG", quality=q, optimize=True)
    return "data:image/jpeg;base64," + _b64.b64encode(buf.getvalue()).decode("ascii")


def _local_to_data_url(rel_or_abs_path: str, max_w: int = 260, q: int = 78) -> str | None:
    """로컬 PNG 경로 → JPEG data URL."""
    if not _PIL_OK:
        return None
    if rel_or_abs_path in _DATA_URL_CACHE:
        return _DATA_URL_CACHE[rel_or_abs_path]
    p = Path(rel_or_abs_path)
    if not p.is_absolute():
        p = Path(r"C:\Users\AD0903\brand_crawler") / rel_or_abs_path
    if not p.exists():
        return None
    try:
        img = _PILImage.open(p)
        url = _pil_to_data_url(img, max_w, q)
        _DATA_URL_CACHE[rel_or_abs_path] = url
        return url
    except Exception:
        return None


def _fetch_with_retry(url: str, timeout: int = 30, retries: int = 2) -> bytes | None:
    """3회 시도 (원본 + 재시도 2회), 브랜드별 맞춤 헤더 사용."""
    import time as _t
    # 브랜드별 Referer 추출 (401/403 방지)
    referer = "/".join(url.split("/")[:3]) + "/"
    headers_list = [
        {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0 Safari/537.36",
            "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
            "Referer": referer,
        },
        {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15",
            "Accept": "*/*",
            "Referer": referer,
        },
        {
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/119.0 Safari/537.36",
            "Accept": "image/*",
        },
    ]
    for i in range(retries + 1):
        try:
            r = _httpx.get(url, headers=headers_list[i % len(headers_list)],
                          timeout=timeout, follow_redirects=True, verify=False)
            if r.status_code == 200 and r.content and len(r.content) > 100:
                return r.content
        except Exception:
            pass
        if i < retries:
            _t.sleep(0.5 * (i + 1))
    return None


def _remote_to_data_url(url: str, max_w: int = 260, q: int = 78, timeout: int = 30) -> str:
    """원격 URL → 다운로드 + JPEG data URL (디스크 캐시, 3회 재시도)."""
    if not _PIL_OK or not _HTTPX_OK:
        return _PLACEHOLDER
    if url in _DATA_URL_CACHE:
        return _DATA_URL_CACHE[url]

    _IMAGE_CACHE_DIR.mkdir(exist_ok=True)
    key = _hashlib.md5(url.encode()).hexdigest()
    cached_jpg = _IMAGE_CACHE_DIR / f"{key}.jpg"

    try:
        if cached_jpg.exists() and cached_jpg.stat().st_size > 100:
            data_bytes = cached_jpg.read_bytes()
        else:
            content = _fetch_with_retry(url, timeout=timeout, retries=2)
            if not content:
                _FAILED_URLS.append(url)
                _DATA_URL_CACHE[url] = _PLACEHOLDER
                return _PLACEHOLDER
            img = _PILImage.open(_io.BytesIO(content))
            if img.mode in ("RGBA", "LA"):
                bg = _PILImage.new("RGB", img.size, (255, 255, 255))
                bg.paste(img, mask=img.split()[-1])
                img = bg
            else:
                img = img.convert("RGB")
            if img.width > max_w:
                ratio = max_w / img.width
                img = img.resize((max_w, int(img.height * ratio)), _PILImage.LANCZOS)
            img.save(cached_jpg, format="JPEG", quality=q, optimize=True)
            data_bytes = cached_jpg.read_bytes()

        data_url = "data:image/jpeg;base64," + _b64.b64encode(data_bytes).decode("ascii")
        _DATA_URL_CACHE[url] = data_url
        return data_url
    except Exception:
        _FAILED_URLS.append(url)
        _DATA_URL_CACHE[url] = _PLACEHOLDER
        return _PLACEHOLDER


def cache_all_images_to_data_url(items: list, max_workers: int = 20) -> None:
    """모든 remote URL을 병렬 다운로드하여 data URL로 치환 (in-place).
    로컬 파일(stills)은 이미 변환돼 있으므로 skip.
    부가 효과: 원본 URL을 `image_url_hires`로 보존 (모달 팝업에서 원본 이미지 표시용)."""
    # 원본 URL 보존 (로더에서 이미 data URL로 치환된 로컬 스틸은 image_url_hires가 별도 설정됨)
    for it in items:
        u = it.get("image_url")
        if isinstance(u, str) and u.startswith(("http://", "https://")) and not it.get("image_url_hires"):
            it["image_url_hires"] = u
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import time as _t

    remote_urls = sorted({
        it["image_url"] for it in items
        if isinstance(it.get("image_url"), str)
        and it["image_url"].startswith(("http://", "https://"))
    })
    if not remote_urls:
        return

    _IMAGE_CACHE_DIR.mkdir(exist_ok=True)
    # 캐시 상태
    cached_count = sum(
        1 for u in remote_urls
        if (_IMAGE_CACHE_DIR / f"{_hashlib.md5(u.encode()).hexdigest()}.jpg").exists()
    )
    to_fetch = len(remote_urls) - cached_count
    print(f"[image cache] {len(remote_urls)}개 원격 URL · 캐시됨 {cached_count} · 다운로드 필요 {to_fetch}")

    t0 = _t.time()
    url_to_data: dict[str, str] = {}
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        fut_to_url = {ex.submit(_remote_to_data_url, u): u for u in remote_urls}
        done = 0
        for fut in as_completed(fut_to_url):
            u = fut_to_url[fut]
            url_to_data[u] = fut.result()
            done += 1
            if to_fetch > 50 and done % 50 == 0:
                print(f"  [{done}/{len(remote_urls)}] ({_t.time()-t0:.1f}s)")

    for it in items:
        u = it.get("image_url")
        if isinstance(u, str) and u in url_to_data:
            it["image_url"] = url_to_data[u]
    # 실패 로그 저장
    if _FAILED_URLS:
        fail_log = _IMAGE_CACHE_DIR / "failed_urls.log"
        fail_log.write_text("\n".join(sorted(set(_FAILED_URLS))), encoding="utf-8")
        print(f"[image cache] 완료 ({_t.time()-t0:.1f}s) — 실패 {len(set(_FAILED_URLS))}건 → {fail_log}")
    else:
        print(f"[image cache] 완료 ({_t.time()-t0:.1f}s) — 모든 이미지 성공")

ROOT = Path(r"C:\Users\AD0903\brand_crawler")       # 입력(raw 데이터·캐시) 위치
SCRIPT_DIR = Path(__file__).resolve().parent        # 출력(HTML·xlsx) 위치
OUT = SCRIPT_DIR / "index.html"                     # GitHub Pages 루트 진입용
FIT_OVERRIDES_XLSX = SCRIPT_DIR / "fit_overrides.xlsx"  # 사용자 수기 입력 반영 (있으면 적용)

# ── Supabase 협업 연동 설정 ────────────────────────────────────────────
# 1) https://supabase.com 프로젝트 > Settings > API 에서 URL·anon key 확인
# 2) anon key는 공개되어도 안전 (RLS 정책으로 보호됨)
# 3) 빈 값이면 localStorage 전용 모드로 폴백 (협업 비활성)
# 4) 환경변수로 주입하거나 아래 상수에 직접 입력
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY", "")
ALLOWED_EMAIL_DOMAIN = "@fnfcorp.com"

# ── 브랜드 ↔ 라인 매핑 ──────────────────────────────────────────────────────────
LINES = [
    ("classic",  "클래식",    "#2E7D32",  ["lacoste", "rl", "celine", "loropiana", "sportyandrich"]),
    ("athleisure","어슬레져",  "#1E40AF",  ["alo", "lululemon", "skims", "miumiu", "prada"]),
    ("sport",    "스포츠",    "#DC2626",  ["wilson", "descente"]),
]
BRAND_LABEL = {
    "alo": "Alo", "wilson": "Wilson", "lacoste": "Lacoste", "rl": "Ralph Lauren",
    "descente": "Descente", "lululemon": "Lululemon",
    "celine": "Celine", "loropiana": "Loro Piana",
    "skims": "Skims", "miumiu": "Miu Miu", "prada": "Prada",
    "sportyandrich": "Sporty & Rich",
}

# ── 핏 3개 열 (기본: 상의·기타) ─────────────────────────────────────────────
FIT_COLS = [
    ("slim",    "슬림핏"),
    ("regular", "레귤러핏"),
    ("over",    "세미오버핏~오버핏"),
]

# 스커트 전용 4개 열 (실루엣·플리츠 조합 기준)
SKIRT_FIT_COLS = [
    ("fitted",     "피티드"),       # 플리츠 없고 · 몸에 딱 붙는 실루엣 (pencil/bodycon/slim)
    ("flare",      "플레어"),       # 플리츠 없고 · 널널한 실루엣 (A-line/flare/full)
    ("pleats",     "플리츠"),       # 플리츠 有 (잔플리츠 포함)
    ("widepleats", "와이드플리츠"),  # 플리츠 간격이 넓은 스커트
]

# 팬츠 및 쇼츠 전용 3개 열 (실루엣 기준, 모든 바지 커버)
BOTTOM_FIT_COLS = [
    ("slim",     "슬림핏"),
    ("straight", "스트레이트핏"),
    ("wide",     "세미와이드~와이드핏"),
]

# 드레스 전용 3개 열 (기장 기준)
DRESS_FIT_COLS = [
    ("mini", "미니"),
    ("midi", "미디"),
    ("maxi", "맥시"),
]

# 탭별 핏 컬럼 (없으면 기본 FIT_COLS 사용)
TAB_FIT_COLS = {
    "skirt": SKIRT_FIT_COLS,
    "bottom": BOTTOM_FIT_COLS,
    "dress": DRESS_FIT_COLS,
}

def fit_cols_for(tab_key: str) -> list[tuple[str, str]]:
    return TAB_FIT_COLS.get(tab_key, FIT_COLS)

# ── 탭 정의: (key, label, 허용 소분류 집합) ─────────────────────────────────
TABS = [
    ("outer",   "아우터",            {"자켓", "가디건"}),
    ("polo",    "폴로",              {"__POLO__"}),
    ("tee",     "티셔츠 및 롱슬리브",  {"반팔티셔츠", "롱슬리브"}),
    ("dress",   "드레스",            {"드레스"}),
    ("skirt",   "스커트",            {"스커트"}),
    ("sweat",   "맨투맨 및 후디",      {"스웻셔츠", "후디"}),
    ("sweater", "스웨터",            {"스웨터"}),
    ("bottom",  "팬츠 및 쇼츠",       {"팬츠", "쇼츠", "레깅스"}),
]

# ── 소분류 정규화 (브랜드별 표기 차이 흡수) ──────────────────────────────────
SUB_NORMALIZE = {
    # 아우터
    "자켓": "자켓", "재킷": "자켓", "블레이저": "자켓", "아우터웨어": "자켓",
    "집업": "자켓", "베스트": "자켓",
    "가디건": "가디건",
    # 상의
    "반팔티셔츠": "반팔티셔츠", "티셔츠 & 탑": "반팔티셔츠", "탑": "반팔티셔츠",
    "롱슬리브": "롱슬리브",
    "스웻셔츠": "스웻셔츠", "후디 & 스웨트셔츠": "스웻셔츠",
    "후디": "후디", "후드": "후디",
    "스웨터": "스웨터",
    "셔츠": "셔츠", "셔츠 & 블라우스": "셔츠",
    "브라탑": "브라탑",
    "폴로 셔츠": "__POLO__",
    "여성 슬림핏 폴로": "__POLO__",
    "여성 레귤러핏 폴로": "__POLO__",
    # 하의
    "팬츠": "팬츠", "데님": "팬츠",
    "쇼츠": "쇼츠",
    "레깅스": "레깅스",
    "스커트": "스커트",
    "드레스": "드레스",
}

POLO_EXCLUDED = {
    "스웨터", "가디건", "스웻셔츠", "후디", "드레스",
    "팬츠", "쇼츠", "레깅스", "스커트", "자켓",
}


# ── 파일 탐색 헬퍼 ───────────────────────────────────────────────────────────
def latest(pattern: str) -> str | None:
    files = sorted(glob(str(ROOT / pattern)))
    return files[-1] if files else None


# ── 소분류 오버라이드 로드 (fabric_overrides.xlsx) ────────────────────────────
_BRAND_KEY_MAP = {
    "Alo": "alo", "Wilson": "wilson", "Lacoste": "lacoste",
    "Ralph Lauren": "rl", "RL": "rl",
    "Descente": "descente", "Lululemon": "lululemon",
    "Celine": "celine", "Loro Piana": "loropiana", "Loropiana": "loropiana",
    "Skims": "skims", "SKIMS": "skims",
    "Miu Miu": "miumiu", "MiuMiu": "miumiu", "Miumiu": "miumiu",
    "Prada": "prada", "PRADA": "prada",
    "Sporty & Rich": "sportyandrich", "Sporty and Rich": "sportyandrich",
    "SportyAndRich": "sportyandrich", "Sportyandrich": "sportyandrich",
}

def load_sub_overrides() -> dict[tuple[str, str], str]:
    """fabric_overrides.xlsx 에서 (brand, 상품명) → 오버라이드된 소분류 매핑.
    generate_combined.py 의 load_overrides()와 동일 포맷 (row1=안내, row2=헤더)."""
    path = ROOT / "fabric_overrides.xlsx"
    if not path.exists():
        return {}
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {}
    h = [str(x).strip() if x else "" for x in rows[1]]
    try:
        bi = h.index("브랜드"); si = h.index("소분류"); ni = h.index("상품명")
    except ValueError:
        return {}
    out = {}
    for r in rows[2:]:
        if not r or not r[bi] or not r[ni]:
            continue
        brand = _BRAND_KEY_MAP.get(str(r[bi]).strip(), str(r[bi]).strip().lower())
        name = str(r[ni]).strip()
        sub = str(r[si] or "").strip()
        if sub:
            out[(brand, name)] = sub
    return out


_SUB_OVERRIDES_CACHE: dict | None = None

def get_sub_overrides() -> dict[tuple[str, str], str]:
    global _SUB_OVERRIDES_CACHE
    if _SUB_OVERRIDES_CACHE is None:
        _SUB_OVERRIDES_CACHE = load_sub_overrides()
    return _SUB_OVERRIDES_CACHE


# ── 소스 파일 로드 ───────────────────────────────────────────────────────────
def load_alo() -> list[dict]:
    path = latest("alo_crawler/alo_yoga_new_arrivals_*_vision.xlsx")
    items = _load_xlsx_kor(path, "alo")
    # Alo AI 누끼 캐시 적용 (handle 매칭) + data URL 변환 (캡처 CORS 회피)
    cache_path = ROOT / "alo_crawler" / "alo_still_images.json"
    if cache_path.exists():
        import re
        still = json.loads(cache_path.read_text(encoding="utf-8"))
        for it in items:
            m = re.search(r"/products/([^/?#]+)", it["product_url"])
            if m and m.group(1) in still:
                p = still[m.group(1)].get("still_path")
                if p:
                    abs_p = Path(p) if Path(p).is_absolute() else (ROOT / p)
                    if abs_p.exists():
                        data_url = _local_to_data_url(str(abs_p).replace("\\", "/"))
                        if data_url:
                            it["image_url"] = data_url
                        else:
                            it["image_url"] = str(abs_p).replace("\\", "/")
                        # 모달용 원본 경로 — file:// URL
                        it["image_url_hires"] = "file:///" + str(abs_p).replace("\\", "/")
    return items


def load_wilson() -> list[dict]:
    path = latest("wilson_crawler/wilson_new_arrivals_*_vision.xlsx")
    items = _load_xlsx_kor(path, "wilson")
    # Wilson 스틸컷 캐시 적용 (handle 매칭)
    cache_path = ROOT / "wilson_crawler" / "wilson_still_images.json"
    if cache_path.exists():
        import re
        still = json.loads(cache_path.read_text(encoding="utf-8"))
        for it in items:
            m = re.search(r"/products/([^/?#]+)", it["product_url"])
            if m and m.group(1) in still:
                url = still[m.group(1)].get("still_url")
                if url:
                    it["image_url"] = url
    return items


def _load_xlsx_kor(path: str, brand: str) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x).strip() if x else "" for x in rows[0]]
    idx = {c: h.index(c) for c in h if c}
    sub_ov = get_sub_overrides()
    items = []
    for r in rows[1:]:
        if (r[idx["대분류"]] or "") != "의류":
            continue
        name = str(r[idx["상품명"]] or "").strip()
        raw_sub = str(r[idx["소분류"]] or "").strip()
        sub = sub_ov.get((brand, name), raw_sub)
        items.append({
            "brand": brand,
            "name": name,
            "sub_raw": sub,
            "fit_raw": str(r[idx.get("핏", -1)] or "").strip() if idx.get("핏", -1) >= 0 else "",
            "pleats_raw": str(r[idx.get("주름", -1)] or "").strip() if idx.get("주름", -1) >= 0 else "",
            "silhouette_raw": str(r[idx.get("실루엣", -1)] or "").strip() if idx.get("실루엣", -1) >= 0 else "",
            "length_raw": str(r[idx.get("기장", -1)] or "").strip() if idx.get("기장", -1) >= 0 else "",
            "image_url": str(r[idx["대표 이미지 URL"]] or "").strip(),
            "product_url": str(r[idx["상품 URL"]] or "").strip(),
        })
    return items


def load_lacoste() -> list[dict]:
    import re as _re
    path = latest("lacoste_crawler/lacoste_new_arrivals_*_vision.xlsx")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x).strip() if x else "" for x in rows[0]]
    idx = {c: h.index(c) for c in h if c}
    sub_ov = get_sub_overrides()
    items = []
    for r in rows[1:]:
        if (r[idx["category_main"]] or "") != "의류":
            continue
        sub_raw = str(r[idx["category_sub"]] or "").strip()
        vision_sub = str(r[idx["vision_sub"]] or "").strip()
        img_url = str(r[idx["image_url"]] or "").strip()
        # Lacoste 스틸컷: suffix _20 → _24 (상품만, 모델 없음)
        if img_url:
            img_url = _re.sub(r"_(\d{2})\.jpg", "_24.jpg", img_url, count=1)
        name = str(r[idx["product_name"]] or "").strip()
        effective = sub_raw or vision_sub
        sub = sub_ov.get(("lacoste", name), effective)
        items.append({
            "brand": "lacoste",
            "name": name,
            "sub_raw": sub,
            "fit_raw": str(r[idx["VISION_FIT"]] or "").strip(),
            "pleats_raw": str(r[idx.get("VISION_PLEATS", -1)] or "").strip() if idx.get("VISION_PLEATS", -1) >= 0 else "",
            "silhouette_raw": str(r[idx.get("VISION_SILHOUETTE", -1)] or "").strip() if idx.get("VISION_SILHOUETTE", -1) >= 0 else "",
            "length_raw": str(r[idx.get("VISION_LENGTH", -1)] or "").strip() if idx.get("VISION_LENGTH", -1) >= 0 else "",
            "image_url": img_url,
            "product_url": str(r[idx["product_url"]] or "").strip(),
        })
    return items


def load_descente() -> list[dict]:
    """Descente: 스포츠 라인. MAIN_IMAGE_URL이 `_R01.JPG`(후면) 이면 N01(전면)로 치환."""
    import re as _re
    path = latest("descente_crawler/descente_new_arrivals_*_vision.xlsx")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x).strip() if x else "" for x in rows[0]]
    idx = {c: h.index(c) for c in h if c}
    sub_ov = get_sub_overrides()
    items = []
    for r in rows[1:]:
        if str(r[idx["CATEGORY_MAIN"]] or "").strip() != "의류":
            continue
        name = str(r[idx["PRODUCT_NAME"]] or "").strip()
        raw_sub = str(r[idx["CATEGORY_SUB"]] or "").strip()
        sub = sub_ov.get(("descente", name), raw_sub)
        img_url = str(r[idx["MAIN_IMAGE_URL"]] or "").strip()
        # R01(후면) → N01(전면) 치환: `/R/` 서브폴더 제거 + `_R0N.JPG` → `_N01.JPG`
        if img_url and "_R01.JPG" in img_url:
            m = _re.search(r"(/product/S/R[12]/[A-Z0-9]+/)R/", img_url)
            if m:
                img_url = img_url.replace(m.group(1) + "R/", m.group(1))
            img_url = _re.sub(r"_R\d{2}\.JPG", "_N01.JPG", img_url)
        items.append({
            "brand": "descente",
            "name": name,
            "sub_raw": sub,
            "fit_raw": str(r[idx.get("VISION_FIT", -1)] or "").strip() if idx.get("VISION_FIT", -1) >= 0 else "",
            "pleats_raw": str(r[idx.get("VISION_PLEATS", -1)] or "").strip() if idx.get("VISION_PLEATS", -1) >= 0 else "",
            "silhouette_raw": str(r[idx.get("VISION_SILHOUETTE", -1)] or "").strip() if idx.get("VISION_SILHOUETTE", -1) >= 0 else "",
            "length_raw": str(r[idx.get("VISION_LENGTH", -1)] or "").strip() if idx.get("VISION_LENGTH", -1) >= 0 else "",
            "image_url": img_url,
            "product_url": str(r[idx["MAIN_PRODUCT_URL"]] or "").strip(),
        })
    return items


def load_lululemon() -> list[dict]:
    """Lululemon: 어슬레져 라인. AI 누끼 캐시(handle 매칭) 적용."""
    import re as _re
    path = latest("lululemon_crawler/lululemon_new_arrivals_*_vision.xlsx")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x).strip() if x else "" for x in rows[0]]
    idx = {c: h.index(c) for c in h if c}
    sub_ov = get_sub_overrides()
    items = []
    for r in rows[1:]:
        if str(r[idx["CATEGORY_MAIN"]] or "").strip() != "의류":
            continue
        name = str(r[idx["PRODUCT_NAME"]] or "").strip()
        raw_sub = str(r[idx["CATEGORY_SUB"]] or "").strip()
        sub = sub_ov.get(("lululemon", name), raw_sub)
        items.append({
            "brand": "lululemon",
            "name": name,
            "sub_raw": sub,
            "fit_raw": str(r[idx.get("VISION_FIT", -1)] or "").strip() if idx.get("VISION_FIT", -1) >= 0 else "",
            "pleats_raw": str(r[idx.get("VISION_PLEATS", -1)] or "").strip() if idx.get("VISION_PLEATS", -1) >= 0 else "",
            "silhouette_raw": str(r[idx.get("VISION_SILHOUETTE", -1)] or "").strip() if idx.get("VISION_SILHOUETTE", -1) >= 0 else "",
            "length_raw": str(r[idx.get("VISION_LENGTH", -1)] or "").strip() if idx.get("VISION_LENGTH", -1) >= 0 else "",
            "image_url": str(r[idx["IMAGE_URL"]] or "").strip(),
            "product_url": str(r[idx["PRODUCT_URL"]] or "").strip(),
        })
    # AI 누끼 캐시 적용 + data URL 변환 (캡처 CORS 회피)
    cache_path = ROOT / "lululemon_crawler" / "lululemon_still_images.json"
    if cache_path.exists():
        still = json.loads(cache_path.read_text(encoding="utf-8"))
        for it in items:
            m = _re.search(r"/([^/]+)\.html", it["product_url"])
            if m and m.group(1) in still:
                p = still[m.group(1)].get("still_path")
                if p:
                    abs_p = Path(p) if Path(p).is_absolute() else (ROOT / p)
                    if abs_p.exists():
                        data_url = _local_to_data_url(str(abs_p).replace("\\", "/"))
                        if data_url:
                            it["image_url"] = data_url
                        else:
                            it["image_url"] = str(abs_p).replace("\\", "/")
                        it["image_url_hires"] = "file:///" + str(abs_p).replace("\\", "/")
    return items


def load_celine() -> list[dict]:
    """Celine: 클래식 라인. IMAGE_URL 컬럼 깨짐 → IMAGE_URL_HIRES 사용."""
    path = latest("celine_crawler/celine_new_arrivals_*_vision.xlsx")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x).strip() if x else "" for x in rows[0]]
    idx = {c: h.index(c) for c in h if c}
    sub_ov = get_sub_overrides()
    items = []
    for r in rows[1:]:
        if str(r[idx["CATEGORY_MAIN"]] or "").strip() != "의류":
            continue
        name = str(r[idx["PRODUCT_NAME"]] or "").strip()
        raw_sub = str(r[idx["CATEGORY_SUB"]] or "").strip()
        sub = sub_ov.get(("celine", name), raw_sub)
        items.append({
            "brand": "celine",
            "name": name,
            "sub_raw": sub,
            "fit_raw": str(r[idx.get("VISION_FIT", -1)] or "").strip() if idx.get("VISION_FIT", -1) >= 0 else "",
            "pleats_raw": str(r[idx.get("VISION_PLEATS", -1)] or "").strip() if idx.get("VISION_PLEATS", -1) >= 0 else "",
            "silhouette_raw": str(r[idx.get("VISION_SILHOUETTE", -1)] or "").strip() if idx.get("VISION_SILHOUETTE", -1) >= 0 else "",
            "length_raw": str(r[idx.get("VISION_LENGTH", -1)] or "").strip() if idx.get("VISION_LENGTH", -1) >= 0 else "",
            "image_url": str(r[idx["IMAGE_URL_HIRES"]] or "").strip(),
            "product_url": str(r[idx["PRODUCT_URL"]] or "").strip(),
        })
    return items


def load_loropiana() -> list[dict]:
    """Loro Piana: 클래식 라인. IMAGE_URL 그대로 사용 (상품컷 위주)."""
    path = latest("loropiana_crawler/loropiana_women_ss_*_vision.xlsx")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x).strip() if x else "" for x in rows[0]]
    idx = {c: h.index(c) for c in h if c}
    sub_ov = get_sub_overrides()
    items = []
    for r in rows[1:]:
        if str(r[idx["CATEGORY_MAIN"]] or "").strip() != "의류":
            continue
        name = str(r[idx["PRODUCT_NAME"]] or "").strip()
        raw_sub = str(r[idx["CATEGORY_SUB"]] or "").strip()
        sub = sub_ov.get(("loropiana", name), raw_sub)
        items.append({
            "brand": "loropiana",
            "name": name,
            "sub_raw": sub,
            "fit_raw": str(r[idx.get("VISION_FIT", -1)] or "").strip() if idx.get("VISION_FIT", -1) >= 0 else "",
            "pleats_raw": str(r[idx.get("VISION_PLEATS", -1)] or "").strip() if idx.get("VISION_PLEATS", -1) >= 0 else "",
            "silhouette_raw": str(r[idx.get("VISION_SILHOUETTE", -1)] or "").strip() if idx.get("VISION_SILHOUETTE", -1) >= 0 else "",
            "length_raw": str(r[idx.get("VISION_LENGTH", -1)] or "").strip() if idx.get("VISION_LENGTH", -1) >= 0 else "",
            "image_url": str(r[idx["IMAGE_URL"]] or "").strip(),
            "product_url": str(r[idx["PRODUCT_URL"]] or "").strip(),
        })
    return items


def load_rl() -> list[dict]:
    path = latest("ralphlauren_crawler/rl_products_vision_*.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    sub_ov = get_sub_overrides()
    items = []
    for p in data:
        if (p.get("category_main") or "") != "의류":
            continue
        imgs = p.get("images") or []
        # RL 스틸컷 우선: URL에 "lifestyle" 포함된 첫 이미지 (모델 없는 상품컷)
        # 없으면 첫 이미지로 폴백
        still = next((u for u in imgs if isinstance(u, str) and "lifestyle" in u), "")
        image_url = still or (imgs[0] if isinstance(imgs, list) and imgs else "")
        name = str(p.get("product_name") or "").strip()
        raw_sub = str(p.get("category_sub") or "").strip()
        sub = sub_ov.get(("rl", name), raw_sub)
        items.append({
            "brand": "rl",
            "name": name,
            "sub_raw": sub,
            "fit_raw": str(p.get("vision_fit") or "").strip(),
            "pleats_raw": str(p.get("vision_pleats") or "").strip(),
            "silhouette_raw": str(p.get("vision_silhouette") or "").strip(),
            "length_raw": str(p.get("vision_length") or "").strip(),
            "image_url": image_url,
            "product_url": str(p.get("product_url") or "").strip(),
        })
    return items


# ── Skims 모델컷 제외 목록 ──
# 사용자 확인(2026-04-20): 다음 5건은 모델 착용 이미지 (Shopify 다른 URL 네이밍 체계)
# 상품컷 대체 URL 없음 → 매트릭스에서 제외
_SKIMS_EXCLUDE_BY_NAME = {
    "MILKY SHEER BOATNECK MINI DRESS",
    "MILKY SHEER MOCK NECK LONG SLEEVE LONG DRESS",
    "COTTON RIB SCOOP NECK HENLEY",
    "FITS EVERYBODY WRAP SUPER CROPPED T-SHIRT",
    "MILKY SHEER MOCK NECK LONG SLEEVE BODYSUIT",
}


def load_skims() -> list[dict]:
    """Skims: 어슬레져 라인. Shopify 한글 헤더 xlsx.
    사용자 확인된 모델컷 5건은 제외."""
    path = latest("skims_crawler/skims_new_arrivals_*_vision.xlsx")
    items = _load_xlsx_kor(path, "skims")
    return [it for it in items if it["name"].strip() not in _SKIMS_EXCLUDE_BY_NAME]


def load_sportyandrich() -> list[dict]:
    """Sporty & Rich: 클래식 라인. Shopify 한글 헤더 xlsx."""
    path = latest("sportyandrich_crawler/sportyandrich_new_arrivals_*_vision.xlsx")
    if not path:
        # _vision suffix 없는 최신본 폴백
        path = latest("sportyandrich_crawler/sportyandrich_new_arrivals_*.xlsx")
    return _load_xlsx_kor(path, "sportyandrich") if path else []


def load_miumiu() -> list[dict]:
    """Miu Miu: 어슬레져 라인. 영문 대문자 헤더(Descente/Lululemon 포맷)."""
    path = latest("miumiu_crawler/miumiu_ready_to_wear_*_vision.xlsx")
    if not path:
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x).strip() if x else "" for x in rows[0]]
    idx = {c: h.index(c) for c in h if c}
    sub_ov = get_sub_overrides()
    items = []
    for r in rows[1:]:
        if str(r[idx["CATEGORY_MAIN"]] or "").strip() != "의류":
            continue
        name = str(r[idx["PRODUCT_NAME"]] or "").strip()
        raw_sub = str(r[idx["CATEGORY_SUB"]] or "").strip()
        sub = sub_ov.get(("miumiu", name), raw_sub)
        items.append({
            "brand": "miumiu",
            "name": name,
            "sub_raw": sub,
            "fit_raw": str(r[idx.get("VISION_FIT", -1)] or "").strip() if idx.get("VISION_FIT", -1) >= 0 else "",
            "pleats_raw": str(r[idx.get("VISION_PLEATS", -1)] or "").strip() if idx.get("VISION_PLEATS", -1) >= 0 else "",
            "silhouette_raw": str(r[idx.get("VISION_SILHOUETTE", -1)] or "").strip() if idx.get("VISION_SILHOUETTE", -1) >= 0 else "",
            "length_raw": str(r[idx.get("VISION_LENGTH", -1)] or "").strip() if idx.get("VISION_LENGTH", -1) >= 0 else "",
            "image_url": str(r[idx["IMAGE_URL"]] or "").strip(),
            "product_url": str(r[idx["PRODUCT_URL"]] or "").strip(),
        })
    return items


def load_prada() -> list[dict]:
    """Prada: 어슬레져 라인. 영문 대문자 헤더. 의류만 필터 (가방/액세서리 제외)."""
    path = latest("prada_crawler/prada_new_in_*_vision.xlsx")
    if not path:
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x).strip() if x else "" for x in rows[0]]
    idx = {c: h.index(c) for c in h if c}
    sub_ov = get_sub_overrides()
    items = []
    for r in rows[1:]:
        if str(r[idx["CATEGORY_MAIN"]] or "").strip() != "의류":
            continue
        name = str(r[idx["PRODUCT_NAME"]] or "").strip()
        raw_sub = str(r[idx["CATEGORY_SUB"]] or "").strip()
        sub = sub_ov.get(("prada", name), raw_sub)
        items.append({
            "brand": "prada",
            "name": name,
            "sub_raw": sub,
            "fit_raw": str(r[idx.get("VISION_FIT", -1)] or "").strip() if idx.get("VISION_FIT", -1) >= 0 else "",
            "pleats_raw": str(r[idx.get("VISION_PLEATS", -1)] or "").strip() if idx.get("VISION_PLEATS", -1) >= 0 else "",
            "silhouette_raw": str(r[idx.get("VISION_SILHOUETTE", -1)] or "").strip() if idx.get("VISION_SILHOUETTE", -1) >= 0 else "",
            "length_raw": str(r[idx.get("VISION_LENGTH", -1)] or "").strip() if idx.get("VISION_LENGTH", -1) >= 0 else "",
            "image_url": str(r[idx["IMAGE_URL"]] or "").strip(),
            "product_url": str(r[idx["PRODUCT_URL"]] or "").strip(),
        })
    return items


# ── 수기 핏 오버라이드 로드 (있으면 적용) ──────────────────────────────────
def load_fit_overrides() -> dict[tuple[str, str], str]:
    """(brand, product_url) → 새 핏값 (슬림핏/레귤러핏/세미오버핏~오버핏)"""
    if not FIT_OVERRIDES_XLSX.exists():
        return {}
    wb = load_workbook(str(FIT_OVERRIDES_XLSX), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x).strip() if x else "" for x in rows[1]]  # row2 = 헤더 (row1 = 안내)
    try:
        bi = h.index("브랜드"); ui = h.index("상품 URL"); ni = h.index("새 핏 ★")
    except ValueError:
        return {}
    brand_key = {"Alo": "alo", "Wilson": "wilson", "Lacoste": "lacoste",
                 "Ralph Lauren": "rl", "RL": "rl"}
    out = {}
    for r in rows[2:]:
        if not r or not r[bi]:
            continue
        b = brand_key.get(str(r[bi]).strip())
        url = str(r[ui] or "").strip()
        new_fit = str(r[ni] or "").strip()
        if b and url and new_fit:
            out[(b, url)] = new_fit
    return out


# ── 분류 로직 ────────────────────────────────────────────────────────────────
def normalize_sub(sub_raw: str) -> str:
    return SUB_NORMALIZE.get(sub_raw, sub_raw)


# 상품명 끝단 기반 제품 타입 추론 — 소분류 오분류 교정용
# 상품명 마지막 어절의 키워드가 강한 시그널 (예: "... 트랙 팬츠" → 팬츠)
_NAME_TYPE_END_KEYWORDS = [
    # (tab_key, 키워드 튜플) — 우선순위 순 (pants/shorts/skirt > outerwear)
    ("bottom", ("팬츠", "pants", "pant", "trouser", "trousers", "바지", "데님", "진")),
    ("bottom", ("쇼츠", "shorts", "short", "반바지")),
    ("skirt", ("스커트", "skirt", "skort")),
    ("outer", ("재킷", "자켓", "jacket", "블레이저", "blazer", "코트", "coat")),
    ("sweater", ("스웨터", "sweater", "니트", "knit")),
    ("sweat", ("후디", "hoodie", "스웻셔츠", "sweatshirt", "맨투맨")),
]

def _infer_tab_from_name(name: str) -> str | None:
    """상품명 토큰을 뒤에서부터 스캔하여 제품 타입 유추.
    마지막에 나오는 키워드가 승리 (예: 트랙 재킷 vs 트랙 팬츠 구분)."""
    if not name:
        return None
    # Tokenize Korean + English (분리자: 공백·괄호·별표·하이픈 등)
    import re as _re
    tokens = _re.split(r"[\s\*\[\]\(\)\-_,./]+", name)
    tokens = [t for t in tokens if t]
    # 뒤에서부터 첫 매칭 키워드 사용
    for tok in reversed(tokens):
        tok_low = tok.lower()
        for tab_key, keywords in _NAME_TYPE_END_KEYWORDS:
            if tok in keywords or tok_low in keywords:
                return tab_key
    return None


def classify_tab(item: dict) -> str | None:
    """상품 → 탭 key. 우선순위: 폴로 하이브리드 > 상품명 강한 시그널 > 소분류 매핑."""
    name = item["name"]
    name_low = name.lower()
    sub_raw = item["sub_raw"]
    norm = normalize_sub(sub_raw)

    # 1) 폴로 하이브리드 (최우선)
    if "폴로" in sub_raw or norm == "__POLO__":
        return "polo"
    if ("폴로" in name or "polo" in name_low) and norm not in POLO_EXCLUDED:
        return "polo"

    # 2) 상품명 끝단 기반 강한 시그널 — 소분류와 불일치 시 name 우선
    #    예: sub=자켓인데 이름이 "...와이드 레그 트랙 팬츠" → bottom 탭으로 이동
    name_tab = _infer_tab_from_name(name)
    if name_tab:
        # 소분류가 이미 name이 가리키는 tab에 해당되면 그대로 사용
        # 불일치하면 name 기반으로 교정
        return name_tab

    # 3) 소분류 → 탭 매핑 (기본 플로우)
    for key, _, subs in TABS:
        if key == "polo":
            continue
        if norm in subs:
            return key
    return None


def classify_fit(item: dict, overrides: dict, tab: str | None = None) -> str | None:
    """상품 → 핏 key. 탭별로 다른 분류:
    - skirt: 플리츠 (논/잔/와이드)
    - bottom: 실루엣 (슬림/스트레이트/세미와이드~와이드)
    - 그 외: 슬림/레귤러/세미오버~오버
    """
    # 스커트 전용 — 플리츠 기반 분류
    if tab == "skirt":
        return _classify_skirt_pleat(item, overrides)

    # 팬츠 및 쇼츠 전용 — 실루엣 기반 분류 (모든 바지 커버)
    if tab == "bottom":
        return _classify_pants_silhouette(item, overrides)

    # 드레스 전용 — 기장 기반 분류 (미니/미디/맥시)
    if tab == "dress":
        return _classify_dress_length(item, overrides)

    # ── 기존: 슬림 / 레귤러 / 세미오버~오버 ─────────────────────────────
    # 1) 사용자 수기 오버라이드
    ov = overrides.get((item["brand"], item["product_url"]))
    if ov:
        if "슬림" in ov: return "slim"
        if "레귤러" in ov: return "regular"
        if "세미" in ov or "오버" in ov: return "over"
    # 2) 슬림 키워드 (상품명 또는 소분류)
    name = item["name"]
    name_low = name.lower()
    if "슬림" in name or "slim" in name_low or "슬림" in item["sub_raw"]:
        return "slim"
    # 3) Vision 핏값
    fit = item["fit_raw"]
    if fit == "Regular":
        return "regular"
    if fit in ("Semi-Over", "Over"):
        return "over"
    return None


def _classify_skirt_pleat(item: dict, overrides: dict) -> str | None:
    """스커트 → 4개 분류 key (fitted/flare/pleats/widepleats).
    규칙:
      - 와이드플리츠 Vision → widepleats
      - 잔플리츠/언발란스 Vision → pleats (합쳐서 "플리츠")
      - 플리츠 없음 → 실루엣·핏·상품명으로 피티드 vs 플레어 판별
    우선순위: 오버라이드 > Vision pleats > (Vision silhouette + 상품명 키워드)"""
    # 1) 사용자 수기 오버라이드
    ov = overrides.get((item["brand"], item["product_url"]))
    if ov:
        if "와이드플리츠" in ov or "언발란스" in ov:
            return "widepleats"
        if "플리츠" in ov:  # 잔플리츠 포함
            return "pleats"
        if "피티드" in ov or "피티트" in ov:
            return "fitted"
        if "플레어" in ov:
            return "flare"

    # 2) Vision 주름값 기반
    p = (item.get("pleats_raw") or "").strip()
    if p in ("와이드플리츠", "언발란스플리츠"):
        return "widepleats"
    if p == "잔플리츠":
        return "pleats"

    # 3) 플리츠 없음 → Vision 실루엣으로 피티드(H라인) vs 플레어(A라인·플레어) 판별
    # Vision 표준값: H라인(피티드 · pencil/straight) / A라인·플레어(플레어 · loose/volume)
    sil_raw = (item.get("silhouette_raw") or "").strip()
    sil = sil_raw.lower()
    fit = (item.get("fit_raw") or "").strip()
    name = (item["name"] or "").lower()

    # 피티드 판정: Vision H라인 OR pencil/bodycon/tight 키워드 OR Slim 핏
    if sil_raw == "H라인":
        return "fitted"
    if any(k in name for k in
           ("pencil", "bodycon", "tight", "fitted", "bodyfit",
            "펜슬", "바디콘", "타이트", "피티드")):
        return "fitted"
    if fit in ("Slim", "Active Slim"):
        return "fitted"

    # 플레어 판정: Vision A라인/플레어 OR flare/a-line/full 키워드 OR Regular/Over 핏
    if sil_raw in ("A라인", "플레어"):
        return "flare"
    if any(k in name for k in
           ("flare", "a-line", "a line", "full skirt",
            "플레어", "에이라인", "볼륨")):
        return "flare"
    if fit in ("Regular", "Semi-Over", "Over"):
        return "flare"

    # 4) 기본값: 플레어 (일반적인 미분류 스커트는 대부분 A-line 계열)
    return "flare"


# 팬츠 fit 분류 키워드
_PANTS_SLIM_KW = ("슬림", "타이트", "스키니", "테이퍼드",
                  "slim", "skinny", "tight", "tapered",
                  "legging", "bike", "cycling")
_PANTS_WIDE_KW = ("와이드", "루즈", "배기", "플레어",
                  "wide", "baggy", "loose", "relaxed", "flare", "flared",
                  "palazzo", "wide leg", "wide-leg", "trouser",
                  "bootcut", "boot-cut", "flared")


def _classify_dress_length(item: dict, overrides: dict) -> str | None:
    """드레스 → 기장 key (mini/midi/maxi).
    우선순위: 오버라이드 > Vision 기장 > 상품명 키워드 > 기본값(미니)."""
    # 1) 오버라이드
    ov = overrides.get((item["brand"], item["product_url"]))
    if ov:
        if "미니" in ov: return "mini"
        if "미디" in ov: return "midi"
        if "맥시" in ov: return "maxi"
    # 2) Vision 기장
    length = (item.get("length_raw") or "").strip()
    if length == "미니": return "mini"
    if length == "미디": return "midi"
    if length == "맥시": return "maxi"
    # 3) 상품명 키워드
    name = (item["name"] or "").lower()
    if any(k in name for k in ("mini", "short", "크롭", "미니")):
        return "mini"
    if any(k in name for k in ("maxi", "long", "맥시", "롱드레스")):
        return "maxi"
    if any(k in name for k in ("midi", "mid", "미디", "knee")):
        return "midi"
    # 4) 기본값 (대부분 크롭이 아니면 미디)
    return "midi"


def _classify_pants_silhouette(item: dict, overrides: dict) -> str:
    """팬츠·쇼츠·레깅스·데님 → 실루엣 key (slim/straight/wide).
    모든 바지가 3개 중 하나에 반드시 분류되도록 설계.
    우선순위: 오버라이드 > Vision 실루엣 > 레깅스 자동 > 이름 키워드 > 기본값(straight)."""
    # 1) 오버라이드
    ov = overrides.get((item["brand"], item["product_url"]))
    if ov:
        if "슬림" in ov: return "slim"
        if "스트레이트" in ov: return "straight"
        if "와이드" in ov: return "wide"
    # 2) Vision 실루엣
    sil = (item.get("silhouette_raw") or "").strip()
    if sil == "테이퍼드":
        return "slim"
    if sil == "스트레이트":
        return "straight"
    if sil == "와이드":
        return "wide"
    # 3) 레깅스 → 무조건 슬림
    if item["sub_raw"] == "레깅스":
        return "slim"
    # 4) 이름 키워드
    name = item["name"] or ""
    name_low = name.lower()
    # 와이드 키워드 우선 (스포츠 팬츠가 slim keyword도 같이 가지는 경우 대비 차후 조정)
    if any(k in name for k in _PANTS_WIDE_KW) or any(k in name_low for k in _PANTS_WIDE_KW):
        return "wide"
    if any(k in name for k in _PANTS_SLIM_KW) or any(k in name_low for k in _PANTS_SLIM_KW):
        return "slim"
    # 5) 기본값: 스트레이트핏 (가장 일반적인 핏)
    return "straight"


def classify_line(brand: str) -> str | None:
    for key, _, _, brands in LINES:
        if brand in brands:
            return key
    return None


# ── ST(SERGIO TACCHINI) 2단계 판정 — core(⭐ 직접 부합) / adapt(✨ 타키니화 가능) ──
# 참조: st_brand_voice_v0.md §1(금지영역·타키니다움), §2(시그니쳐 계보), §2.4(경쟁사 분류 4구분)
# 목표: 100~200건 선별. 이원화 마크로 후속 작업 우선순위 부여.
#
# 반환값:
#   "core"  → ⭐ ST 직접 부합 (현 5개 시그니쳐 계보에 바로 얹을 수 있음)
#   "adapt" → ✨ 타키니화 가능 (실루엣·디자인 참조 가치, 소재/디테일 수정 후 사용)
#   None    → 부적합
#
# 절대 배제: 가운·이브닝·코르셋·란제리·슬립드레스·자수/시퀸 과다, 브라탑 단독, Prada 로고 플레이
#
# ⭐ CORE (직접 부합) — 시그니쳐 계보 기반 강한 매칭:
#   - 폴로 → 플레잉폴로 (Lacoste/Sporty&Rich/Wilson만 · 테니스 헤리티지 3사)
#   - 스커트 → 플라잉스커트 (테니스 헤리티지 3사 OR 플리츠 有)
#   - 쇼츠 → 에센셜쇼츠 (테니스 헤리티지 3사만)
#   - 아우터 → 쿠쉬라이트 (키워드: crop/track/windbreak/bomber)
#   - 맨투맨 → Italian Heritage 클럽웨어 (SR/Lacoste + retro/track/classic/italia 키워드)
#
# ✨ ADAPT (타키니화 가능) — 실루엣 참조용, 소재·디테일 ST 기준으로 전환:
#   - 럭셔리 4사(Miu Miu/Prada/Celine/Loro Piana) 자켓 (블레이저/봄버/크롭/재킷 키워드)
#   - 럭셔리 4사 플리츠 스커트 (Vision pleats = 잔/와이드/언발란스)
#   - Skims 쇼츠 (Body-Lined 실루엣 참조 → Cotton Blend Jersey로 전환)
#   - Skims 플리츠 스커트
def classify_st_match(item: dict, tab: str) -> str | None:
    brand = item["brand"]
    name = (item["name"] or "").lower()
    sub = item["sub_raw"] or ""
    pleats = (item.get("pleats_raw") or "").strip()

    # ── 절대 배제선 ───────────────────────────────────────────────
    if any(k in name for k in
           ("gown", "evening", "corset", "lingerie", "slip dress",
            "가운", "이브닝", "코르셋",
            "sequin", "embroidered", "embellish",
            "시퀸", "스팽글", "자수")):
        return None
    if sub == "브라탑":
        return None
    # Prada 삼각 로고 플레이 과다
    if brand == "prada" and any(k in name for k in ("triangle", "logo", "삼각")):
        return None

    # 플리츠 있는 스커트 (잔플리츠·와이드플리츠·언발란스) — 플라잉스커트 계보 참조용
    has_pleats = pleats in ("잔플리츠", "와이드플리츠", "언발란스플리츠")

    # ST-aligned 키워드 (테니스·클래식 헤리티지 · 이탈리안 시그널)
    st_kw = any(k in name for k in (
        "tennis", "court", "track", "pleat", "classic", "essential",
        "italia", "italian", "riviera", "heritage", "club", "yacht",
        "yale", "princeton", "varsity", "college", "ivy", "oxford",
        "crew", "pique", "piqué", "harrington",
        "테니스", "코트", "트랙", "클래식", "플리츠", "헤리티지", "클럽"))

    length = (item.get("length_raw") or "").strip()

    # ── ✨ ADAPT: 럭셔리 4사 ─────────────────────────────────────
    #   자켓은 특정 실루엣 키워드(블레이저/봄버/크롭/바시티/트랙/해링턴)로 좁힘
    #   (단순 "재킷/jacket" 키워드는 거의 모든 럭셔리 아우터에 있어서 제외)
    if brand in ("miumiu", "prada", "celine", "loropiana"):
        if tab == "outer" and any(k in name for k in
            ("blazer", "bomber", "crop", "varsity", "track", "harrington",
             "블레이저", "봄버", "크롭", "바시티", "트랙", "해링턴")):
            return "adapt"
        # 플리츠 스커트만 (플라잉스커트 spin-off 참조)
        if tab == "skirt" and has_pleats:
            return "adapt"
        # 드레스: 미니만 (§6.8 Cotton Blend Jersey × Mini Dress 실험 방향)
        if tab == "dress" and length == "미니":
            return "adapt"
        return None

    # ── ✨ ADAPT: Skims (쇼츠 + 플리츠 스커트 + 미니 드레스 · Body-Lined 참조) ──
    if brand == "skims":
        if tab == "bottom" and "쇼츠" in sub:
            return "adapt"
        if tab == "skirt" and has_pleats:
            return "adapt"
        if tab == "dress" and length == "미니":
            return "adapt"
        return None

    # ── ⭐ CORE: 브랜드 그룹 ─────────────────────────────────────
    # TH2 = 순수 테니스 브랜드 (Lacoste, Wilson) — 무조건 매칭
    # SR  = Sporty & Rich — 헤리티지 스포츠 브랜드, 키워드 필터 적용 (볼륨 큼)
    th2 = brand in ("lacoste", "wilson")
    is_sr = brand == "sportyandrich"

    # 폴로 (플레잉폴로 계보)
    if tab == "polo":
        if th2:
            return "core"
        if is_sr and st_kw:
            return "core"
        return None

    # 스커트 (플라잉스커트 계보)
    if tab == "skirt":
        if th2:
            return "core"
        if is_sr and (st_kw or has_pleats):
            return "core"
        if has_pleats:  # 다른 브랜드도 플리츠 있으면 core
            return "core"
        return None

    # 쇼츠 (에센셜쇼츠 계보)
    if tab == "bottom":
        is_shorts = "쇼츠" in sub
        if is_shorts and th2:
            return "core"
        if is_shorts and is_sr and st_kw:
            return "core"
        return None

    # 아우터 (쿠쉬라이트 경량 우븐 · 특정 실루엣 키워드만)
    if tab == "outer":
        kushlight_kw = any(k in name for k in
            ("track", "crop", "windbreak", "wind-break", "bomber",
             "트랙", "크롭", "바람막이", "봄버"))
        if kushlight_kw:
            return "core"
        return None

    # 드레스 (§6.8 Cotton Blend Jersey × Active Slim × Mini Dress 실험 방향)
    # ⭐: 테니스 헤리티지 브랜드(Lacoste/SR/Wilson/RL)의 미니·미디 드레스 — 폴로/테니스 드레스
    # ⭐: 그 외 브랜드도 tennis/polo/court/classic 키워드 있으면 포함
    if tab == "dress":
        if length == "맥시":
            return None  # 맥시는 ST 방향과 거리 있음
        if th2 or is_sr or brand == "rl":
            if length in ("미니", "미디"):
                return "core"
            return None
        # 다른 브랜드(Alo/Lulu/Descente) — tennis/polo 키워드 + 미니만
        if length == "미니" and any(k in name for k in
               ("tennis", "polo", "court", "classic", "athletic",
                "테니스", "폴로", "코트", "클래식")):
            return "core"
        # Alo/Lulu 미니 드레스 (키워드 없어도) — Body-Lined athleisure 참조
        if brand in ("alo", "lululemon") and length == "미니":
            return "adapt"
        return None

    # 맨투맨 (French Terry 클럽웨어 · SR/Lacoste + 레트로 키워드)
    if tab == "sweat":
        if brand in ("sportyandrich", "lacoste"):
            if any(k in name for k in
                   ("track", "classic", "varsity", "college", "retro",
                    "italia", "riviera", "club", "heritage",
                    "트랙", "클래식", "레트로")):
                return "core"
        return None

    # 티셔츠·스웨터 → 전면 제외 (시그니쳐 미확립 + RA5 배제 원칙)
    return None


# ── 빌드: tab → line → fit → [items] ─────────────────────────────────────────
def build_matrix(items: list[dict], overrides: dict) -> dict:
    matrix = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    # 색상 변형 URL이 다른 같은 상품명은 1개만 매트릭스에 노출
    # 우선 이미지가 data URL(로컬 스틸컷)인 항목을 앞으로 정렬하여 색상 중 가장 좋은 이미지 선택
    items_sorted = sorted(
        items,
        key=lambda it: (0 if str(it.get("image_url", "")).startswith("data:") else 1,),
    )
    seen: set[tuple[str, str]] = set()
    for it in items_sorted:
        # dedup 키: (브랜드, 상품명 정규화)
        dedup_key = (it["brand"], (it["name"] or "").strip().lower())
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        tab = classify_tab(it)
        if not tab:
            continue
        line = classify_line(it["brand"])
        if not line:
            continue
        fit = classify_fit(it, overrides, tab=tab)
        if not fit:
            continue
        it["st_match"] = classify_st_match(it, tab)  # "core" / "adapt" / None
        it["tab"] = tab
        it["line"] = line
        it["fit"] = fit
        matrix[tab][line][fit].append(it)
    return matrix


# ── HTML 렌더링 ──────────────────────────────────────────────────────────────
LINE_LABEL_MAP = {k: lbl for k, lbl, _, _ in LINES}


def render_html(matrix: dict) -> str:
    tabs_nav = "\n".join(
        f'<button class="tab-btn{" active" if i == 0 else ""}" data-tab="{k}">'
        f'<span class="tab-lbl">{html.escape(lbl)}</span>'
        f'</button>'
        for i, (k, lbl, _) in enumerate(TABS)
    )
    # 선택된 항목 탭 추가 (오른쪽)
    tabs_nav += (
        '\n<button class="tab-btn tab-btn-selected" data-tab="__selected__">'
        '<span class="tab-lbl">✓ MD PICK <span class="tab-count" id="tab-count-selected">0</span></span>'
        '</button>'
    )
    # 디자이너 PICK 탭 추가 (맨 오른쪽)
    tabs_nav += (
        '\n<button class="tab-btn tab-btn-pick" data-tab="__pick__">'
        '<span class="tab-lbl">🎨 디자이너 PICK <span class="tab-count" id="tab-count-pick">0</span></span>'
        '</button>'
    )

    # 각 상품의 공유 가능한 고유 ID (product_url 우선, 없으면 brand+name 해시)
    import hashlib as _h
    def _item_id(it):
        base = (it.get("product_url") or "").strip() or (it["brand"] + "|" + (it.get("name") or ""))
        return _h.md5(base.encode("utf-8")).hexdigest()[:16]

    # JS에서 사용할 전체 아이템 사전 (id → metadata). 선택 탭/백업/엑셀용.
    items_js_map = {}

    panels = []
    for i, (tab_key, tab_lbl, _) in enumerate(TABS):
        cols = fit_cols_for(tab_key)
        axis_kicker = "PLEAT" if tab_key == "skirt" else "FIT"
        rows_html = []
        for line_key, line_lbl, color, _ in LINES:
            cells = []
            for fit_key, fit_lbl in cols:
                items = matrix[tab_key][line_key][fit_key]
                thumb_parts = []
                for it in items:
                    iid = _item_id(it)
                    match = it.get("st_match")  # "core" / "adapt" / None
                    badge = ""
                    if match == "core":
                        badge = '<span class="badge badge-core" title="ST 직접 부합 (플레잉폴로·플라잉스커트·쿠쉬라이트 계보)">⭐</span>'
                    elif match == "adapt":
                        badge = '<span class="badge badge-adapt" title="타키니화 가능 (실루엣 참조 · 소재/디테일 수정 필요)">✨</span>'
                    # 모달·백업용 전체 메타데이터 (JS 측에서 id로 조회)
                    items_js_map[iid] = {
                        "id": iid,
                        "brand": it["brand"],
                        "brand_label": BRAND_LABEL.get(it["brand"], it["brand"]),
                        "name": it["name"],
                        "sub": it.get("sub_raw") or "",
                        "tab": tab_key,
                        "tab_label": tab_lbl,
                        "line": line_key,
                        "line_label": LINE_LABEL_MAP[line_key],
                        "fit": fit_key,
                        "fit_label": fit_lbl,
                        "image_thumb": it["image_url"],
                        "image_hires": it.get("image_url_hires") or it["image_url"],
                        "product_url": it.get("product_url") or "",
                        "match": match or "",
                    }
                    thumb_parts.append(
                        f'<div class="thumb" data-id="{iid}" data-match="{match or ""}" role="button" tabindex="0">'
                        f'<img src="{html.escape(it["image_url"])}" alt="" loading="lazy">'
                        f'{badge}'
                        f'<label class="chk" onclick="event.stopPropagation()">'
                        f'<input type="checkbox" class="item-chk" data-id="{iid}"></label>'
                        f'</div>'
                    )
                thumbs = "".join(thumb_parts)
                empty_cls = " is-empty" if not items else ""
                cells.append(
                    f'<td class="gallery-cell{empty_cls}">'
                    f'<div class="count-pill">{len(items)}</div>'
                    f'<div class="gallery">{thumbs}</div></td>'
                )
            rows_html.append(
                f'<tr style="--line-color:{color}">'
                f'<th class="line-th">'
                f'<div class="line-th-inner">'
                f'<span class="line-dot"></span>'
                f'<span class="line-label">{html.escape(line_lbl)}</span>'
                f'</div>'
                f'</th>'
                + "".join(cells) + "</tr>"
            )
        fit_headers = "".join(
            f'<th class="fit-th"><span class="fit-kicker">{axis_kicker}</span>'
            f'<span class="fit-lbl">{html.escape(lbl)}</span></th>'
            for _, lbl in cols
        )
        panel_desc = "LINE × PLEAT" if tab_key == "skirt" else "LINE × FIT"
        panels.append(
            f'<div class="panel{" active" if i == 0 else ""}" data-panel="{tab_key}">'
            f'<div class="panel-head">'
            f'<div class="panel-meta">'
            f'<span class="panel-kicker">CATEGORY</span>'
            f'<h2 class="panel-title">{html.escape(tab_lbl)}</h2>'
            f'<span class="panel-desc">{panel_desc}</span>'
            f'</div>'
            f'<div class="panel-actions">'
            f'<button class="btn-ghost" data-action="copy">'
            f'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect width="14" height="14" x="8" y="8" rx="2" ry="2"></rect><path d="M4 16c-1.1 0-2-.9-2-2V4c0-1.1.9-2 2-2h10c1.1 0 2 .9 2 2"></path></svg>'
            f'<span>클립보드 복사</span></button>'
            f'<button class="btn-solid" data-action="download">'
            f'<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path><polyline points="7 10 12 15 17 10"></polyline><line x1="12" y1="15" x2="12" y2="3"></line></svg>'
            f'<span>이미지 저장</span></button>'
            f'</div>'
            f'</div>'
            f'<div class="matrix-scroll">'
            f'<table class="matrix"><thead><tr>'
            f'<th class="corner-th"></th>{fit_headers}'
            f'</tr></thead>'
            f'<tbody>{"".join(rows_html)}</tbody></table>'
            f'</div>'
            f'</div>'
        )

    total = sum(
        len(matrix[t][l][f])
        for t, _, _ in TABS
        for l, _, _, _ in LINES
        for f, _ in fit_cols_for(t)
    )

    # 라인별 브랜드 및 상품 수 (매트릭스에 실제 노출되는 건 기준)
    line_items: dict[str, dict[str, int]] = {lk: {} for lk, _, _, _ in LINES}
    for t_k, _, _ in TABS:
        for l_k, _, _, _ in LINES:
            for f_k, _ in fit_cols_for(t_k):
                for it in matrix[t_k][l_k][f_k]:
                    line_items[l_k][it["brand"]] = line_items[l_k].get(it["brand"], 0) + 1

    # 라인 카드 (대시보드 카드 스타일)
    line_cards = []
    for idx, (line_key, line_lbl, color, brand_keys) in enumerate(LINES, 1):
        brand_pills = "".join(
            f'<div class="brand-pill">'
            f'<span class="bp-name">{html.escape(BRAND_LABEL[b])}</span>'
            f'<span class="bp-count">{line_items[line_key].get(b, 0)}</span>'
            f'</div>'
            for b in brand_keys
        )
        total_n = sum(line_items[line_key].values())
        line_cards.append(
            f'<div class="line-card" style="--line-color:{color}">'
            f'<div class="lc-head">'
            f'<div class="lc-meta">'
            f'<span class="lc-kicker">LINE {idx:02d}</span>'
            f'<span class="lc-name">{html.escape(line_lbl)}</span>'
            f'</div>'
            f'<div class="lc-count">'
            f'<span class="lc-n">{total_n}</span>'
            f'<span class="lc-n-lbl">ITEMS</span>'
            f'</div>'
            f'</div>'
            f'<div class="lc-brands">{brand_pills}</div>'
            f'</div>'
        )

    brand_n = sum(1 for _, _, _, bks in LINES for _ in bks)
    hero_stats = f"""
<div class="hero-stats">
  <div class="stat"><div class="stat-n">{total}</div><div class="stat-lbl">Classified Items</div></div>
  <div class="stat"><div class="stat-n">{brand_n}</div><div class="stat-lbl">Brands</div></div>
  <div class="stat"><div class="stat-n">{len(TABS)}</div><div class="stat-lbl">Categories</div></div>
  <div class="stat"><div class="stat-n">{len(LINES)}×3</div><div class="stat-lbl">Line × Fit</div></div>
</div>"""

    # items_js_map → JSON (JS에서 사용). data URL이 크므로 압축 없이 덤프.
    items_json = json.dumps(items_js_map, ensure_ascii=False)
    tab_order_json = json.dumps([k for k, _, _ in TABS])
    core_count = sum(1 for v in items_js_map.values() if v["match"] == "core")
    adapt_count = sum(1 for v in items_js_map.values() if v["match"] == "adapt")

    # Supabase 협업 설정 (JS에 주입)
    supabase_url_json = json.dumps(SUPABASE_URL or "")
    supabase_key_json = json.dumps(SUPABASE_ANON_KEY or "")
    supabase_domain_json = json.dumps(ALLOWED_EMAIL_DOMAIN)

    return f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="utf-8">
<title>F&amp;F · Brand Matrix Dashboard</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@500;600&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/@supabase/supabase-js@2/dist/umd/supabase.min.js"></script>
<style>
:root {{
    --bg: #f5f6f8;
    --bg-2: #eceef2;
    --surface: #ffffff;
    --surface-soft: #fbfbfc;
    --ink: #0a0a0b;
    --ink-2: #3a3d45;
    --ink-3: #767a84;
    --ink-4: #a5a9b2;
    --line: #e5e7ec;
    --line-2: #d9dce2;
    --line-strong: #15171c;
    --accent: #0a0a0b;
    --classic: #047857;
    --athleisure: #2563eb;
    --sport: #dc2626;
    --radius: 14px;
    --radius-sm: 8px;
    --shadow-sm: 0 1px 2px rgba(10,12,18,.04);
    --shadow: 0 1px 3px rgba(10,12,18,.05), 0 1px 2px rgba(10,12,18,.03);
    --shadow-lg: 0 10px 30px -10px rgba(10,12,18,.18), 0 2px 6px rgba(10,12,18,.05);
}}
* {{ box-sizing: border-box; }}
html, body {{ height: 100%; }}
body {{
    margin: 0; padding: 0;
    font-family: 'Inter','Apple SD Gothic Neo','Noto Sans KR',system-ui,-apple-system,sans-serif;
    background: var(--bg);
    color: var(--ink);
    font-feature-settings: 'tnum' 1, 'cv11' 1;
    -webkit-font-smoothing: antialiased;
    line-height: 1.5;
}}
.app {{ max-width: 1600px; margin: 0 auto; padding: 28px 32px 72px; }}

/* ── TOP BAR ───────────────────────────── */
.topbar {{
    display: flex; justify-content: space-between; align-items: center;
    padding: 14px 24px; margin-bottom: 24px;
    background: var(--surface);
    border: 1px solid var(--line);
    border-radius: var(--radius);
    box-shadow: var(--shadow-sm);
}}
.topbar-brand {{
    display: flex; align-items: center; gap: 10px;
    font-size: 14px; font-weight: 700;
    letter-spacing: -0.01em; color: var(--ink);
}}
.topbar-brand .logo-dot {{
    width: 8px; height: 8px; border-radius: 2px;
    background: linear-gradient(135deg, var(--classic), var(--athleisure), var(--sport));
}}
.topbar-meta {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px; color: var(--ink-3); letter-spacing: 0.05em;
}}

/* ── HERO ───────────────────────────────── */
.hero {{
    display: grid; grid-template-columns: 1.3fr 2fr; gap: 24px;
    margin-bottom: 28px;
}}
.hero-card {{
    background: var(--surface); padding: 28px 32px;
    border: 1px solid var(--line); border-radius: var(--radius);
    box-shadow: var(--shadow-sm);
}}
.eyebrow {{
    display: inline-flex; align-items: center; gap: 8px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px; font-weight: 600;
    letter-spacing: 0.15em; text-transform: uppercase;
    color: var(--ink-3); padding: 4px 10px;
    background: var(--bg-2); border-radius: 100px;
    margin-bottom: 14px;
}}
.hero-card h1 {{
    margin: 0 0 12px;
    font-weight: 800; font-size: 36px; line-height: 1.1;
    letter-spacing: -0.03em; color: var(--ink);
}}
.hero-desc {{
    margin: 0; font-size: 14px; color: var(--ink-2);
    max-width: 420px; line-height: 1.6;
}}
.hero-stats {{
    display: grid; grid-template-columns: repeat(4, 1fr); gap: 0;
    background: var(--surface); border: 1px solid var(--line);
    border-radius: var(--radius); overflow: hidden;
    box-shadow: var(--shadow-sm);
}}
.stat {{ padding: 24px 22px; border-right: 1px solid var(--line); position: relative; }}
.stat:last-child {{ border-right: 0; }}
.stat-n {{
    font-size: 32px; font-weight: 800; line-height: 1;
    color: var(--ink); letter-spacing: -0.025em;
    font-variant-numeric: tabular-nums;
}}
.stat-lbl {{
    font-size: 11px; font-weight: 500;
    color: var(--ink-3); margin-top: 10px;
    letter-spacing: 0.01em;
}}

/* ── SECTION HEADING ─────────────────────── */
.section-head {{
    display: flex; align-items: baseline; justify-content: space-between;
    margin-bottom: 16px;
}}
.section-title-group {{ display: flex; align-items: baseline; gap: 14px; }}
.section-kicker {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px; font-weight: 600;
    letter-spacing: 0.1em; color: var(--ink-3);
    text-transform: uppercase;
}}
.section-title {{
    margin: 0; font-size: 20px; font-weight: 700;
    letter-spacing: -0.015em; color: var(--ink);
}}

/* ── LINE CARDS ─────────────────────────── */
.lines-wrap {{ margin-bottom: 32px; }}
.line-cards {{
    display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px;
}}
.line-card {{
    background: var(--surface);
    border: 1px solid var(--line);
    border-radius: var(--radius);
    padding: 22px 24px;
    position: relative; overflow: hidden;
    transition: transform .2s ease, box-shadow .2s ease, border-color .2s ease;
    box-shadow: var(--shadow-sm);
}}
.line-card::before {{
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px;
    background: var(--line-color);
}}
.line-card:hover {{
    transform: translateY(-2px);
    box-shadow: var(--shadow-lg);
    border-color: var(--line-2);
}}
.lc-head {{
    display: flex; justify-content: space-between; align-items: flex-start;
    padding-bottom: 16px; margin-bottom: 16px;
    border-bottom: 1px solid var(--line);
}}
.lc-meta {{ display: flex; flex-direction: column; gap: 6px; }}
.lc-kicker {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px; font-weight: 600;
    letter-spacing: 0.15em; color: var(--ink-3);
    text-transform: uppercase;
}}
.lc-name {{
    font-size: 22px; font-weight: 800;
    letter-spacing: -0.02em; color: var(--line-color);
    line-height: 1;
}}
.lc-count {{ display: flex; flex-direction: column; align-items: flex-end; gap: 4px; }}
.lc-n {{
    font-size: 28px; font-weight: 800; line-height: 1;
    color: var(--ink); font-variant-numeric: tabular-nums;
    letter-spacing: -0.02em;
}}
.lc-n-lbl {{
    font-size: 10px; font-weight: 600; color: var(--ink-3);
    letter-spacing: 0.15em; font-family: 'JetBrains Mono', monospace;
}}
.lc-brands {{ display: flex; flex-wrap: wrap; gap: 6px; }}
.brand-pill {{
    display: inline-flex; align-items: center; gap: 6px;
    padding: 5px 8px 5px 12px;
    background: var(--surface-soft);
    border: 1px solid var(--line);
    border-radius: 100px;
    font-size: 12px; transition: background .15s;
}}
.brand-pill:hover {{ background: var(--bg-2); }}
.bp-name {{ font-weight: 500; color: var(--ink-2); }}
.bp-count {{
    display: inline-flex; align-items: center; justify-content: center;
    min-width: 20px; padding: 1px 6px;
    background: var(--line-color); color: white;
    font-size: 10px; font-weight: 700;
    border-radius: 100px;
    font-variant-numeric: tabular-nums;
}}

/* ── LEGEND (ST 마크 설명) ───────────── */
.legend {{
    background: var(--surface);
    border: 1px solid var(--line);
    border-left: 4px solid #fbbf24;
    border-radius: var(--radius);
    padding: 18px 22px;
    margin-bottom: 18px;
    box-shadow: var(--shadow-sm);
}}
.legend-title {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px; font-weight: 700;
    letter-spacing: 0.12em; color: var(--ink-2);
    text-transform: uppercase;
    margin-bottom: 14px;
}}
.legend-items {{
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 18px;
}}
.legend-item {{
    display: flex; align-items: flex-start; gap: 12px;
    padding: 10px 12px;
    background: var(--surface-soft);
    border: 1px solid var(--line);
    border-radius: var(--radius-sm);
}}
.legend-item .badge {{
    flex-shrink: 0;
    font-size: 14px; padding: 4px 8px;
    border-radius: 4px;
    font-weight: 700;
}}
.legend-item .badge-core {{ color: #b45309; background: #fef3c7; border: 1px solid #fbbf24; }}
.legend-item .badge-adapt {{ color: #6b21a8; background: #f3e8ff; border: 1px solid #c084fc; }}
.legend-item .badge-placeholder {{
    flex-shrink: 0;
    width: 28px; height: 28px; display: inline-flex;
    align-items: center; justify-content: center;
    color: var(--ink-4); font-weight: 700;
    background: var(--bg-2); border: 1px solid var(--line);
    border-radius: 4px;
}}
.legend-item-mute {{ opacity: .7; }}
.legend-text {{ display: flex; flex-direction: column; gap: 4px; min-width: 0; }}
.legend-text strong {{
    font-size: 13px; font-weight: 700; color: var(--ink);
    letter-spacing: -0.01em;
}}
.legend-text span {{ font-size: 12px; color: var(--ink-3); line-height: 1.45; }}
@media (max-width: 1100px) {{ .legend-items {{ grid-template-columns: 1fr; }} }}

/* ── TABS (modern pill style) ──────────── */
.tabs-wrap {{
    background: var(--surface);
    border: 1px solid var(--line);
    border-radius: var(--radius);
    padding: 6px;
    margin-bottom: 16px;
    box-shadow: var(--shadow-sm);
    display: flex;
    align-items: center;
    gap: 12px;
}}
.tabs {{
    display: flex; gap: 2px;
    flex-wrap: wrap;
    flex: 1; min-width: 0;
}}
.zoom-ctrl {{
    display: flex; align-items: center; gap: 6px;
    padding: 0 8px 0 10px;
    flex-shrink: 0;
    border-left: 1px solid var(--line);
    height: 32px;
}}
.zoom-ctrl .zoom-icon {{
    font-size: 14px; line-height: 1;
    color: var(--ink-3);
}}
.zoom-ctrl button {{
    width: 26px; height: 26px;
    border: 1px solid var(--line-2);
    background: var(--surface);
    border-radius: 6px;
    cursor: pointer;
    font-size: 14px; font-weight: 700; line-height: 1;
    color: var(--ink-2);
    display: inline-flex; align-items: center; justify-content: center;
    padding: 0;
}}
.zoom-ctrl button:hover {{ background: var(--bg-2); color: var(--ink); }}
.zoom-ctrl input[type="range"] {{
    width: 120px; accent-color: #0a0a0b;
    cursor: pointer;
}}
.zoom-ctrl .zoom-val {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px; font-weight: 600;
    color: var(--ink-3);
    min-width: 40px; text-align: right;
    font-variant-numeric: tabular-nums;
}}
.tab-btn {{
    padding: 10px 16px;
    border: 0; background: transparent;
    cursor: pointer; font-family: inherit;
    color: var(--ink-3);
    transition: all .15s ease;
    border-radius: var(--radius-sm);
    font-size: 13px; font-weight: 600;
    letter-spacing: -0.005em;
}}
.tab-btn:hover {{ color: var(--ink); background: var(--bg-2); }}
.tab-btn.active {{
    color: var(--surface); background: var(--ink);
    box-shadow: 0 1px 2px rgba(10,12,18,.2);
}}
.tab-lbl {{ font-size: 13px; font-weight: 600; }}

/* ── PANEL ───────────────────────────────── */
.panel {{ display: none; }}
.panel.active {{ display: block; animation: fadeIn .3s ease; }}
@keyframes fadeIn {{ from {{ opacity: 0; transform: translateY(4px); }} to {{ opacity: 1; transform: none; }} }}
.panel-head {{
    display: flex; justify-content: space-between; align-items: center;
    padding: 18px 22px;
    background: var(--surface);
    border: 1px solid var(--line);
    border-radius: var(--radius) var(--radius) 0 0;
    border-bottom: 0;
}}
.panel-meta {{ display: flex; align-items: baseline; gap: 14px; }}
.panel-kicker {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px; font-weight: 600;
    color: var(--ink-3); letter-spacing: 0.15em;
    text-transform: uppercase;
}}
.panel-title {{
    margin: 0; font-size: 20px; font-weight: 800;
    letter-spacing: -0.02em; color: var(--ink);
}}
.panel-desc {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px; font-weight: 500; color: var(--ink-4);
    letter-spacing: 0.1em;
}}
.panel-actions {{ display: flex; gap: 8px; }}
.btn-ghost, .btn-solid {{
    display: inline-flex; align-items: center; gap: 7px;
    padding: 8px 14px; border-radius: var(--radius-sm);
    font-family: inherit; font-size: 12px; font-weight: 600;
    letter-spacing: -0.005em;
    cursor: pointer;
    transition: all .15s ease;
    border: 1px solid transparent;
}}
.btn-ghost {{
    background: var(--surface);
    border-color: var(--line-2);
    color: var(--ink-2);
}}
.btn-ghost:hover {{ background: var(--bg-2); color: var(--ink); }}
.btn-solid {{
    background: var(--ink); color: var(--surface);
    border-color: var(--ink);
}}
.btn-solid:hover {{ background: #1f2128; }}
.btn-ghost:disabled, .btn-solid:disabled {{
    opacity: .5; cursor: not-allowed;
}}

/* ── MATRIX ───────────────────────────── */
.matrix-scroll {{
    overflow-x: auto;
    background: var(--surface);
    border: 1px solid var(--line);
    border-radius: 0 0 var(--radius) var(--radius);
    box-shadow: var(--shadow-sm);
}}
.matrix-scroll::-webkit-scrollbar {{ height: 10px; }}
.matrix-scroll::-webkit-scrollbar-track {{ background: var(--bg-2); }}
.matrix-scroll::-webkit-scrollbar-thumb {{ background: var(--line-2); border-radius: 5px; }}
.matrix-scroll::-webkit-scrollbar-thumb:hover {{ background: var(--ink-4); }}
.matrix {{
    width: max-content; min-width: 100%;
    border-collapse: separate; border-spacing: 0;
    font-variant-numeric: tabular-nums;
}}
.matrix thead th {{
    background: var(--surface-soft);
    color: var(--ink);
    padding: 14px 18px;
    text-align: left;
    vertical-align: middle;
    border-bottom: 1px solid var(--line);
    position: sticky; top: 0; z-index: 15;
}}
.fit-th {{ min-width: calc(var(--thumb-w, 100px) * 8 + 110px); }}
.fit-kicker {{
    display: block;
    font-family: 'JetBrains Mono', monospace;
    font-size: 9px; font-weight: 600;
    letter-spacing: 0.15em; color: var(--ink-4);
    margin-bottom: 3px; text-transform: uppercase;
}}
.fit-lbl {{
    display: block;
    font-size: 13px; font-weight: 700;
    letter-spacing: -0.01em;
}}
.corner-th {{
    width: 180px; min-width: 180px;
    background: var(--surface-soft) !important;
    position: sticky; left: 0; z-index: 20 !important;
}}
.matrix tbody tr + tr > * {{ border-top: 1px solid var(--line); }}
.line-th {{
    width: 180px; min-width: 180px;
    padding: 22px 20px;
    vertical-align: top;
    background: var(--surface);
    border-right: 1px solid var(--line);
    position: sticky; left: 0; z-index: 10;
}}
.line-th-inner {{
    display: flex; flex-direction: column; gap: 10px;
}}
.line-dot {{
    display: block; width: 12px; height: 12px;
    border-radius: 50%; background: var(--line-color);
    box-shadow: 0 0 0 3px rgba(15, 17, 21, 0.08);
}}
.line-label {{
    font-size: 15px; font-weight: 700;
    letter-spacing: -0.015em; color: var(--ink);
}}
.gallery-cell {{
    padding: 20px 18px 22px;
    vertical-align: top; position: relative;
    border-left: 1px solid var(--line);
    min-width: calc(var(--thumb-w, 100px) * 8 + 110px);
    background: var(--surface);
}}
.gallery-cell.is-empty {{ background: var(--surface-soft); }}
.count-pill {{
    position: absolute; top: 14px; right: 18px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px; font-weight: 700;
    letter-spacing: 0.05em;
    padding: 4px 10px;
    background: var(--ink); color: var(--surface);
    border-radius: 100px;
    font-variant-numeric: tabular-nums;
}}
.gallery-cell.is-empty .count-pill {{
    background: var(--line-2); color: var(--ink-3);
}}
.gallery {{
    display: grid;
    grid-template-columns: repeat(8, var(--thumb-w, 100px));
    gap: 10px;
    margin-top: 26px;
    justify-content: start;
}}
.thumb {{
    position: relative;
    width: var(--thumb-w, 100px);
    height: calc(var(--thumb-w, 100px) * 1.35);
    overflow: hidden;
    border-radius: var(--radius-sm);
    border: 1px solid var(--line);
    background: var(--surface-soft);
    transition: transform .15s ease, box-shadow .15s ease, border-color .15s ease;
    cursor: zoom-in;
}}
.thumb:hover {{
    transform: translateY(-2px);
    box-shadow: var(--shadow);
    border-color: var(--ink-4);
}}
.thumb.is-selected {{
    border-color: #0a0a0b;
    box-shadow: 0 0 0 2px #0a0a0b;
}}
.thumb.is-pick {{
    border-color: #9333ea;
    box-shadow: 0 0 0 2px #9333ea;
}}
.thumb.is-pick.is-selected {{
    border-color: #0a0a0b;
    box-shadow: 0 0 0 2px #0a0a0b, 0 0 0 5px #c084fc;
}}
.thumb img {{
    width: 100%; height: 100%;
    object-fit: contain; object-position: center;
    display: block;
    pointer-events: none;
}}
.thumb .badge {{
    position: absolute; top: 3px; left: 3px;
    font-size: 11px; line-height: 1;
    padding: 2px 5px;
    border-radius: 4px;
    background: rgba(255,255,255,.94);
    border: 1px solid rgba(0,0,0,.06);
    box-shadow: 0 1px 2px rgba(0,0,0,.1);
    pointer-events: auto;
    cursor: help;
    font-weight: 700;
}}
.thumb .badge-core {{ color: #b45309; background: #fef3c7; border-color: #fbbf24; }}
.thumb .badge-adapt {{ color: #6b21a8; background: #f3e8ff; border-color: #c084fc; }}
.thumb .chk {{
    position: absolute; top: 4px; right: 4px;
    display: inline-flex; align-items: center; justify-content: center;
    width: 20px; height: 20px;
    background: rgba(255,255,255,.92);
    border: 1px solid var(--line-2);
    border-radius: 4px;
    cursor: pointer;
    z-index: 2;
}}
.thumb .chk input {{
    width: 14px; height: 14px;
    margin: 0; cursor: pointer;
    accent-color: #0a0a0b;
}}

/* ── MODAL ─────────────────────────────── */
.modal-backdrop {{
    position: fixed; inset: 0;
    background: rgba(10, 12, 18, 0.72);
    backdrop-filter: blur(4px);
    display: none; align-items: center; justify-content: center;
    z-index: 200; padding: 20px;
}}
.modal-backdrop.open {{ display: flex; }}
.modal {{
    background: var(--surface);
    border-radius: var(--radius);
    max-width: 1000px; width: 100%;
    max-height: 92vh;
    display: grid; grid-template-columns: 1.4fr 1fr;
    overflow: hidden;
    box-shadow: 0 30px 60px rgba(0,0,0,.4);
    animation: modalIn .18s ease;
}}
@keyframes modalIn {{
    from {{ opacity: 0; transform: translateY(12px) scale(.98); }}
    to {{ opacity: 1; transform: none; }}
}}
.modal-img-wrap {{
    background: #f5f6f8;
    display: flex; align-items: center; justify-content: center;
    padding: 24px;
    min-height: 300px;
    max-height: 92vh;
    position: relative;
}}
.modal-img-wrap img {{
    max-width: 100%; max-height: 84vh;
    object-fit: contain;
}}
.modal-img-wrap .img-loading {{
    position: absolute;
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px; color: var(--ink-3);
}}
.modal-body {{
    padding: 28px 28px 22px;
    display: flex; flex-direction: column;
    gap: 14px;
    overflow-y: auto;
}}
.modal-kicker {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px; font-weight: 600;
    letter-spacing: 0.15em; color: var(--ink-3);
    text-transform: uppercase;
}}
.modal-title {{
    margin: 0;
    font-size: 20px; font-weight: 700;
    letter-spacing: -0.015em; color: var(--ink);
    line-height: 1.3;
}}
.modal-tags {{
    display: flex; flex-wrap: wrap; gap: 6px;
    margin-top: 2px;
}}
.modal-tag {{
    font-size: 11px; font-weight: 600;
    padding: 4px 10px;
    background: var(--bg-2);
    color: var(--ink-2);
    border-radius: 100px;
    letter-spacing: -0.005em;
}}
.modal-tag.brand {{ background: var(--ink); color: #fff; }}
.modal-tag.star {{ background: #fef3c7; color: #92400e; border: 1px solid #fbbf24; }}
.modal-tag.adapt {{ background: #f3e8ff; color: #6b21a8; border: 1px solid #c084fc; }}
.modal-actions {{
    margin-top: auto;
    display: flex; flex-direction: column; gap: 10px;
}}
.modal-chk-row {{
    display: flex; align-items: center; gap: 10px;
    padding: 12px 14px;
    background: var(--bg-2);
    border-radius: var(--radius-sm);
    cursor: pointer; user-select: none;
}}
.modal-chk-row:hover {{ background: #e5e7ec; }}
.modal-chk-row input {{ width: 18px; height: 18px; accent-color: #0a0a0b; }}
.modal-chk-row .lbl {{ font-size: 13px; font-weight: 600; color: var(--ink); }}
.modal-chk-row.pick {{
    background: #f3e8ff;
    border: 1px solid #c084fc;
}}
.modal-chk-row.pick:hover {{ background: #e9d5ff; }}
.modal-chk-row.pick input {{ accent-color: #9333ea; }}
.modal-chk-row.pick .lbl {{ color: #6b21a8; }}
.modal-link {{
    display: inline-flex; align-items: center; gap: 8px;
    padding: 11px 16px;
    background: var(--ink); color: var(--surface);
    border-radius: var(--radius-sm);
    font-size: 13px; font-weight: 600;
    text-decoration: none;
    transition: background .15s;
}}
.modal-link:hover {{ background: #1f2128; }}
.modal-close {{
    position: absolute; top: 12px; right: 12px;
    width: 36px; height: 36px;
    border: 0; background: rgba(255,255,255,.9);
    border-radius: 50%; cursor: pointer;
    font-size: 18px; color: var(--ink);
    z-index: 3;
    display: flex; align-items: center; justify-content: center;
}}
.modal-close:hover {{ background: #fff; transform: rotate(90deg); }}
.modal-close {{ transition: transform .2s, background .15s; }}

/* ── TOPBAR ACTIONS (선택 카운트 + 버튼) ── */
.topbar-actions {{
    display: flex; align-items: center; gap: 10px;
}}
.sel-count {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 12px; font-weight: 600;
    color: var(--ink-2);
    padding: 6px 12px;
    background: var(--bg-2);
    border-radius: 100px;
    letter-spacing: -0.005em;
}}
.sel-count strong {{ color: var(--ink); font-weight: 800; }}
.topbar-actions button {{
    padding: 7px 13px;
    background: var(--surface);
    border: 1px solid var(--line-2);
    border-radius: var(--radius-sm);
    color: var(--ink-2);
    font-family: inherit; font-size: 12px; font-weight: 600;
    cursor: pointer; transition: all .15s;
}}
.topbar-actions button:hover {{ background: var(--bg-2); color: var(--ink); }}
.topbar-actions button.primary {{
    background: var(--ink); color: var(--surface); border-color: var(--ink);
}}
.topbar-actions button.primary:hover {{ background: #1f2128; }}
.topbar-actions button.danger:hover {{
    border-color: var(--sport); color: var(--sport);
}}
.sel-divider {{
    display: inline-block; width: 1px; height: 18px;
    background: var(--line-2); margin: 0 4px;
}}
#btn-download-rec {{
    background: #fef3c7; border-color: #fbbf24; color: #92400e;
}}
#btn-download-rec:hover {{ background: #fde68a; }}
#btn-backup-hires {{
    background: #fee2e2; border-color: #f87171; color: #991b1b;
    font-weight: 700;
}}
#btn-backup-hires:hover {{ background: #fecaca; }}
#btn-backup-pick, #btn-export-pick-zip {{
    background: #f3e8ff; border-color: #c084fc; color: #6b21a8;
}}
#btn-backup-pick:hover, #btn-export-pick-zip:hover {{ background: #e9d5ff; }}
#btn-clear-pick.danger:hover {{
    border-color: #9333ea; color: #9333ea;
}}

/* 독립 다중 필터: 활성 필터 있으면 기본 숨김 후 해당 카테고리만 재노출 (OR 결합) */
body.has-filter .thumb {{ display: none; }}
body.show-core .thumb[data-match="core"],
body.show-adapt .thumb[data-match="adapt"],
body.show-exc .thumb[data-match=""] {{ display: block; }}
body.has-filter .gallery-cell.hidden-all {{ opacity: .35; }}

/* 레전드 박스 클릭 인터랙션 */
.legend-item {{
    cursor: pointer;
    transition: transform .15s, background .15s, box-shadow .15s;
    outline: none;
}}
.legend-item:hover {{ transform: translateY(-1px); background: var(--bg-2); }}
.legend-item:focus-visible {{ box-shadow: 0 0 0 2px var(--ink); }}
.legend-item[data-filter="core"].active {{
    background: #fef3c7; box-shadow: 0 0 0 2px #fbbf24;
}}
.legend-item[data-filter="adapt"].active {{
    background: #f3e8ff; box-shadow: 0 0 0 2px #c084fc;
}}
.legend-item[data-filter="exc"].active {{
    background: #e5e7eb; box-shadow: 0 0 0 2px #6b7280;
}}
.legend-hint {{
    font-size: 11px; font-weight: 500; color: var(--ink-3);
    letter-spacing: 0;
}}

/* ── SELECTED TAB PANEL ─────────────── */
.tab-btn-selected {{
    margin-left: auto;
    background: #fef3c7 !important;
    color: #92400e !important;
}}
.tab-btn-selected.active {{
    background: #0a0a0b !important;
    color: #fef3c7 !important;
}}
.tab-count {{
    display: inline-block;
    margin-left: 6px; padding: 1px 7px;
    background: rgba(0,0,0,.15);
    border-radius: 100px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px; font-weight: 700;
    font-variant-numeric: tabular-nums;
}}
.tab-btn-selected.active .tab-count {{ background: rgba(255,255,255,.2); }}
.tab-btn-pick {{
    background: #f3e8ff !important;
    color: #6b21a8 !important;
}}
.tab-btn-pick.active {{
    background: #9333ea !important;
    color: #fff !important;
}}
.tab-btn-pick.active .tab-count {{ background: rgba(255,255,255,.2); }}

#panel-selected, #panel-pick {{
    display: none; padding: 24px;
    background: var(--surface); border: 1px solid var(--line);
    border-radius: var(--radius);
    box-shadow: var(--shadow-sm);
}}
#panel-selected.active, #panel-pick.active {{ display: block; animation: fadeIn .3s ease; }}
.sel-empty {{
    padding: 48px 20px; text-align: center;
    color: var(--ink-3); font-size: 14px;
}}
.sel-group {{ margin-bottom: 28px; }}
.sel-group-head {{
    display: flex; align-items: baseline; gap: 12px;
    padding-bottom: 10px; margin-bottom: 14px;
    border-bottom: 1px solid var(--line);
}}
.sel-group-title {{ font-size: 15px; font-weight: 700; letter-spacing: -0.015em; }}
.sel-group-count {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px; color: var(--ink-3);
    font-variant-numeric: tabular-nums;
}}
.sel-gallery {{
    display: grid;
    grid-template-columns: repeat(auto-fill, var(--thumb-w, 100px));
    gap: 10px;
}}

/* ── TOAST ─────────────────────────────── */
#toast {{
    position: fixed; bottom: 32px; left: 50%; transform: translateX(-50%);
    padding: 12px 20px;
    background: var(--ink); color: var(--surface);
    border-radius: 100px;
    font-size: 13px; font-weight: 600;
    letter-spacing: -0.01em;
    box-shadow: 0 20px 40px -12px rgba(0,0,0,.3);
    opacity: 0; pointer-events: none;
    transition: opacity .2s, transform .2s;
    z-index: 100;
}}
#toast.show {{ opacity: 1; transform: translateX(-50%) translateY(-4px); }}
#toast.error {{ background: var(--sport); }}

/* ── FOOTER ─────────────────────────────── */
.footer {{
    margin-top: 48px; padding-top: 20px;
    border-top: 1px solid var(--line);
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px; color: var(--ink-3);
    letter-spacing: 0.05em;
    display: flex; justify-content: space-between;
}}

/* Responsive */
@media (max-width: 1100px) {{
    .hero {{ grid-template-columns: 1fr; }}
    .line-cards {{ grid-template-columns: 1fr; }}
}}
@media (max-width: 720px) {{
    .app {{ padding: 20px 16px 48px; }}
    .hero-card h1 {{ font-size: 28px; }}
    .hero-stats {{ grid-template-columns: repeat(2, 1fr); }}
    .stat {{ border-right: 0; border-bottom: 1px solid var(--line); }}
    .panel-head {{ flex-direction: column; align-items: flex-start; gap: 12px; }}
}}

/* Capture-mode: temporary class added during html2canvas to ensure full gallery renders */
.capture-mode .matrix-scroll {{ overflow: visible; }}
.capture-mode .corner-th, .capture-mode .line-th {{ position: static; }}
.capture-mode .matrix thead th {{ position: static; }}

/* ── Supabase 협업: Auth Modal · User Pill · Attribution Badge ── */
.auth-modal-inner {{
    max-width: 440px; padding: 32px; background: #fff;
    border-radius: 16px; box-shadow: var(--shadow-lg);
    display: flex; flex-direction: column; gap: 14px;
}}
.auth-title {{ font-size: 22px; font-weight: 700; color: var(--ink); }}
.auth-desc {{ color: var(--ink-3); font-size: 13px; line-height: 1.55; margin-bottom: 4px; }}
.auth-row {{ display: flex; gap: 8px; }}
.auth-row input {{
    flex: 1; padding: 10px 12px; border: 1px solid var(--line);
    border-radius: 8px; font: inherit; font-size: 14px;
}}
.auth-row input:focus {{ outline: none; border-color: var(--ink); }}
.auth-row button {{
    padding: 10px 16px; border-radius: 8px; background: var(--ink);
    color: #fff; border: 0; cursor: pointer; font-weight: 600; font-size: 13px;
}}
.auth-row button:disabled {{ opacity: .5; cursor: not-allowed; }}
.auth-status {{ min-height: 18px; font-size: 12.5px; line-height: 1.5; }}
.auth-status.ok {{ color: #047857; }}
.auth-status.error {{ color: #dc2626; }}
.auth-hint {{ font-size: 11.5px; color: var(--ink-4); }}
.auth-skip {{
    align-self: flex-start; background: transparent; border: 1px solid var(--line);
    color: var(--ink-3); padding: 6px 12px; border-radius: 6px;
    cursor: pointer; font-size: 11.5px; font-weight: 500;
}}

.user-pill {{
    position: fixed; top: 12px; right: 12px; z-index: 1000;
    background: var(--ink); color: #fff; border-radius: 100px;
    padding: 6px 10px 6px 14px; font-size: 12px; font-weight: 600;
    display: flex; align-items: center; gap: 8px;
    box-shadow: var(--shadow);
}}
.user-pill .u-dot {{ width: 6px; height: 6px; border-radius: 50%; background: #10b981; }}
.user-pill button {{
    background: transparent; border: 1px solid rgba(255,255,255,.25);
    color: #fff; padding: 2px 8px; border-radius: 10px;
    font-size: 11px; cursor: pointer;
}}
.user-pill button:hover {{ background: rgba(255,255,255,.1); }}

/* 팀원 attribution 배지 (썸네일 좌하단) */
.attr-badge {{
    position: absolute; bottom: 4px; left: 4px;
    background: rgba(15,17,21,.85); color: #fff;
    font-size: 9px; font-weight: 700; letter-spacing: .01em;
    padding: 2px 6px; border-radius: 100px;
    z-index: 3; cursor: help; pointer-events: auto;
    display: flex; align-items: center; gap: 3px;
    line-height: 1.2;
}}
.attr-badge .me-dot {{ width: 5px; height: 5px; border-radius: 50%; background: #10b981; }}
.capture-mode .attr-badge {{ display: none; }}
</style></head>
<body>
<div class="app">
  <div class="topbar">
    <div class="topbar-brand"><span class="logo-dot"></span>F&amp;F · Brand Matrix</div>
    <div class="topbar-actions">
      <button id="btn-download-rec" title="⭐✨ 추천 항목 전체 이미지 ZIP (브라우저 · 일부 썸네일 폴백)">⭐✨ 추천 ZIP</button>
      <button id="btn-backup-hires" class="hires" title="⭐✨ 추천 백업 JSON 저장 후 download_hires.py로 원본 고화질 다운로드">🔥 고화질용 백업</button>
      <span class="sel-divider"></span>
      <span class="sel-count">MD PICK <strong id="sel-count">0</strong> / {total}</span>
      <button id="btn-backup" title="MD PICK 백업 JSON 다운로드">MD 백업</button>
      <button id="btn-export-xlsx" class="primary" title="MD PICK 이미지 ZIP (브라우저)">MD ZIP</button>
      <button id="btn-clear-all" class="danger" title="MD PICK 전체 해제">MD 해제</button>
      <span class="sel-divider"></span>
      <span class="sel-count">PICK <strong id="pick-count">0</strong></span>
      <button id="btn-backup-pick" title="디자이너 PICK 백업 JSON 다운로드">PICK 백업</button>
      <button id="btn-export-pick-zip" title="디자이너 PICK 이미지 ZIP (브라우저)">PICK ZIP</button>
      <button id="btn-clear-pick" class="danger" title="디자이너 PICK 전체 해제">PICK 해제</button>
    </div>
  </div>

  <section class="hero">
    <div class="hero-card">
      <div class="eyebrow">Brand Intelligence · 2026</div>
      <h1>라인 × 핏 매트릭스</h1>
      <p class="hero-desc">
        8개 브랜드를 <strong>라인 · 아이템 카테고리 · 핏</strong> 세 축으로 분류하여
        포지셔닝을 실시간으로 시각화합니다.
      </p>
    </div>{hero_stats}
  </section>

  <section class="lines-wrap">
    <div class="section-head">
      <div class="section-title-group">
        <span class="section-kicker">01</span>
        <h2 class="section-title">라인별 브랜드 구성</h2>
      </div>
    </div>
    <div class="line-cards">{"".join(line_cards)}</div>
  </section>

  <section class="matrix-wrap">
    <div class="section-head">
      <div class="section-title-group">
        <span class="section-kicker">02</span>
        <h2 class="section-title">아이템 카테고리별 라인 × 핏</h2>
      </div>
    </div>

    <div class="legend">
      <div class="legend-title">SERGIO TACCHINI 이미지 재생성 후보 — 마크 가이드 <span class="legend-hint">· 박스 클릭으로 토글 · 복수 선택 가능</span></div>
      <div class="legend-items">
        <div class="legend-item" data-filter="core" role="button" tabindex="0" title="⭐ 직접 부합 토글 (복수 선택 가능)">
          <span class="badge badge-core" style="position:static">⭐</span>
          <div class="legend-text">
            <strong>ST 직접 부합 ({core_count}건)</strong>
            <span>현 시그니쳐 계보(플레잉폴로·플라잉스커트·쿠쉬라이트·에센셜쇼츠)에 바로 얹을 수 있는 아이템</span>
          </div>
        </div>
        <div class="legend-item" data-filter="adapt" role="button" tabindex="0" title="✨ 타키니화 가능 토글 (복수 선택 가능)">
          <span class="badge badge-adapt" style="position:static">✨</span>
          <div class="legend-text">
            <strong>타키니화 가능 ({adapt_count}건)</strong>
            <span>럭셔리 4사의 재킷·플리츠 스커트 또는 Skims의 Body-Lined 쇼츠·스커트 — 실루엣 참조 후 소재·디테일은 ST 기준으로 전환</span>
          </div>
        </div>
        <div class="legend-item legend-item-mute" data-filter="exc" role="button" tabindex="0" title="제외 토글 (복수 선택 가능)">
          <span class="badge-placeholder">—</span>
          <div class="legend-text">
            <strong>제외 ({total - core_count - adapt_count}건)</strong>
            <span>§1 금지영역(가운·이브닝·코르셋·자수 과다·브라탑 단독) 또는 시그니쳐 계보 비매칭</span>
          </div>
        </div>
      </div>
    </div>

    <div class="tabs-wrap">
      <nav class="tabs">{tabs_nav}</nav>
      <div class="zoom-ctrl" title="이미지 썸네일 확대/축소 (갤러리 전용)">
        <span class="zoom-icon">🔍</span>
        <button id="zoom-out" aria-label="축소">−</button>
        <input type="range" id="zoom-slider" min="60" max="280" value="100" step="10" aria-label="이미지 배율">
        <button id="zoom-in" aria-label="확대">＋</button>
        <span class="zoom-val" id="zoom-val">100%</span>
      </div>
    </div>
    {"".join(panels)}

    <div id="panel-selected" data-panel="__selected__">
      <div class="section-head">
        <div class="section-title-group">
          <span class="section-kicker">MD PICK</span>
          <h2 class="section-title">MD PICK <span id="panel-selected-count" style="color:var(--ink-3);font-weight:600">(0)</span></h2>
        </div>
      </div>
      <div id="selected-body"></div>
    </div>

    <div id="panel-pick" data-panel="__pick__">
      <div class="section-head">
        <div class="section-title-group">
          <span class="section-kicker">DESIGNER PICK</span>
          <h2 class="section-title">디자이너 PICK <span id="panel-pick-count" style="color:var(--ink-3);font-weight:600">(0)</span></h2>
        </div>
      </div>
      <div id="pick-body"></div>
    </div>
  </section>

  <footer class="footer">
    <div>F&amp;F · BRAND CRAWLER DASHBOARD</div>
    <div>GENERATED · {total} ITEMS</div>
  </footer>
</div>

<div id="toast"></div>

<!-- User Pill (로그인 시 상단 우측 표시) -->
<div id="user-pill" class="user-pill" style="display:none">
  <span class="u-dot"></span>
  <span id="user-pill-email">—</span>
  <button id="user-pill-logout" title="로그아웃">로그아웃</button>
</div>

<!-- Auth Modal (세션 없을 때 차단) -->
<div id="auth-modal" class="modal-backdrop" aria-hidden="true">
  <div class="auth-modal-inner" role="dialog" aria-labelledby="auth-title">
    <div class="auth-title" id="auth-title">F&amp;F · Brand Matrix 로그인</div>
    <div class="auth-desc">
      팀원과 MD PICK / 디자이너 PICK 을 실시간으로 공유하려면
      <strong>F&amp;F 이메일</strong>로 로그인해주세요. 받은 링크를 클릭하면 자동 로그인됩니다.
    </div>
    <div class="auth-row">
      <input id="auth-email" type="email" placeholder="name@fnfcorp.com" autocomplete="email">
      <button id="auth-send" type="button">매직 링크 전송</button>
    </div>
    <div id="auth-status" class="auth-status"></div>
    <div class="auth-hint">이메일을 확인해 링크를 클릭하면 이 탭이 자동으로 로그인됩니다.</div>
    <button id="auth-skip" class="auth-skip" type="button">읽기 전용으로 보기</button>
  </div>
</div>

<div id="modal" class="modal-backdrop" aria-hidden="true">
  <div class="modal" role="dialog">
    <div class="modal-img-wrap">
      <span class="img-loading" id="modal-loading">이미지 로딩 중...</span>
      <img id="modal-img" alt="">
      <button class="modal-close" id="modal-close" aria-label="닫기">✕</button>
    </div>
    <div class="modal-body">
      <div class="modal-kicker" id="modal-kicker">CATEGORY · LINE · FIT</div>
      <h3 class="modal-title" id="modal-title">—</h3>
      <div class="modal-tags" id="modal-tags"></div>
      <div class="modal-actions">
        <label class="modal-chk-row">
          <input type="checkbox" id="modal-chk">
          <span class="lbl">MD PICK</span>
        </label>
        <label class="modal-chk-row pick">
          <input type="checkbox" id="modal-pick-chk">
          <span class="lbl">🎨 디자이너 PICK</span>
        </label>
        <a class="modal-link" id="modal-link" href="#" target="_blank" rel="noopener">
          상품 상세 페이지 열기 →
        </a>
      </div>
    </div>
  </div>
</div>

<script>
// ── Item metadata map (id → full item) ──
const ITEMS = {items_json};
</script>

<script src="https://cdnjs.cloudflare.com/ajax/libs/jszip/3.10.1/jszip.min.js"></script>
<script>
// ── Supabase 협업 설정 (Python에서 주입) ──
const SUPABASE_CFG = {{
    url: {supabase_url_json},
    key: {supabase_key_json},
    domain: {supabase_domain_json}
}};
const SUPABASE_ENABLED = !!(SUPABASE_CFG.url && SUPABASE_CFG.key);
let sb = null;
let currentUser = null;  // {{ email }} or null (읽기전용/오프라인)
// 팀원 전체 상태: Map<item_id, Set<email>>
const remoteMdPicks = new Map();
const remoteDesignerPicks = new Map();
if (SUPABASE_ENABLED && window.supabase) {{
    sb = window.supabase.createClient(SUPABASE_CFG.url, SUPABASE_CFG.key, {{
        auth: {{ persistSession: true, autoRefreshToken: true, detectSessionInUrl: true }}
    }});
}}

// ── 로컬 상태 (낙관적 업데이트 + Supabase 미설정 시 폴백) ──
const STORAGE_KEY = 'line_matrix_selected_v1';
const PICK_KEY = 'line_matrix_designer_pick_v1';
const TOTAL_ITEMS = {total};
let selected = new Set();
let designerPicks = new Set();
// Supabase 비활성 상태에서만 localStorage 로드 (활성 시 로그인 후 DB에서 로드)
if (!SUPABASE_ENABLED) {{
    try {{
        const raw = localStorage.getItem(STORAGE_KEY);
        if (raw) selected = new Set(JSON.parse(raw));
    }} catch (e) {{ console.warn('localStorage read failed:', e); }}
    try {{
        const raw = localStorage.getItem(PICK_KEY);
        if (raw) designerPicks = new Set(JSON.parse(raw));
    }} catch (e) {{ console.warn('pick storage read failed:', e); }}
}}

function persist() {{
    if (!SUPABASE_ENABLED) {{
        try {{
            localStorage.setItem(STORAGE_KEY, JSON.stringify([...selected]));
        }} catch (e) {{ console.error('localStorage write failed:', e); }}
    }}
    updateSelCount();
}}

function updateSelCount() {{
    const n = selected.size;
    const el = document.getElementById('sel-count');
    if (el) el.textContent = n;
    const tabEl = document.getElementById('tab-count-selected');
    if (tabEl) tabEl.textContent = n;
    const pEl = document.getElementById('panel-selected-count');
    if (pEl) pEl.textContent = '(' + n + ')';
}}

async function applySelected(id, on) {{
    // Supabase 활성 + 미로그인 → 쓰기 차단, 읽기전용 안내
    if (SUPABASE_ENABLED && !currentUser) {{
        showToast('로그인 후 사용할 수 있습니다', true);
        // 체크박스 되돌리기
        document.querySelectorAll('input.item-chk[data-id="' + id + '"]').forEach(el => el.checked = !on);
        const mchk = document.getElementById('modal-chk');
        if (mchk && mchk.dataset.id === id) mchk.checked = !on;
        return;
    }}
    // 낙관적 로컬 업데이트
    if (on) selected.add(id); else selected.delete(id);
    document.querySelectorAll('input.item-chk[data-id="' + id + '"]').forEach(el => {{ el.checked = on; }});
    document.querySelectorAll('.thumb[data-id="' + id + '"]').forEach(el => {{
        el.classList.toggle('is-selected', on);
    }});
    const mchk = document.getElementById('modal-chk');
    if (mchk && mchk.dataset.id === id) mchk.checked = on;
    persist();
    if (document.getElementById('panel-selected').classList.contains('active')) {{
        renderSelectedPanel();
    }}
    // Supabase 동기화
    if (sb && currentUser) {{
        try {{
            if (on) {{
                const {{ error }} = await sb.from('md_picks').upsert(
                    {{ user_email: currentUser.email, item_id: id }},
                    {{ onConflict: 'user_email,item_id' }}
                );
                if (error) throw error;
                if (!remoteMdPicks.has(id)) remoteMdPicks.set(id, new Set());
                remoteMdPicks.get(id).add(currentUser.email);
            }} else {{
                const {{ error }} = await sb.from('md_picks').delete()
                    .eq('user_email', currentUser.email).eq('item_id', id);
                if (error) throw error;
                if (remoteMdPicks.has(id)) {{
                    remoteMdPicks.get(id).delete(currentUser.email);
                    if (!remoteMdPicks.get(id).size) remoteMdPicks.delete(id);
                }}
            }}
            renderAttributionBadgesFor(id);
        }} catch (e) {{
            console.error('[md_picks sync error]', e);
            showToast('동기화 실패 — 네트워크 확인 필요', true);
        }}
    }}
}}

// 페이지 로드 시 체크박스 상태 복원
function restoreAllCheckboxes() {{
    document.querySelectorAll('input.item-chk').forEach(el => {{
        const on = selected.has(el.dataset.id);
        el.checked = on;
        const thumb = el.closest('.thumb');
        if (thumb) thumb.classList.toggle('is-selected', on);
    }});
    updateSelCount();
}}

// ── 디자이너 PICK 상태 관리 ──
function persistPicks() {{
    if (!SUPABASE_ENABLED) {{
        try {{
            localStorage.setItem(PICK_KEY, JSON.stringify([...designerPicks]));
        }} catch (e) {{ console.error('pick persist failed:', e); }}
    }}
    updatePickCount();
}}

function updatePickCount() {{
    const n = designerPicks.size;
    const el = document.getElementById('pick-count');
    if (el) el.textContent = n;
    const tabEl = document.getElementById('tab-count-pick');
    if (tabEl) tabEl.textContent = n;
    const pEl = document.getElementById('panel-pick-count');
    if (pEl) pEl.textContent = '(' + n + ')';
}}

async function applyPick(id, on) {{
    if (SUPABASE_ENABLED && !currentUser) {{
        showToast('로그인 후 사용할 수 있습니다', true);
        const mpchk0 = document.getElementById('modal-pick-chk');
        if (mpchk0 && mpchk0.dataset.id === id) mpchk0.checked = !on;
        return;
    }}
    if (on) designerPicks.add(id); else designerPicks.delete(id);
    document.querySelectorAll('.thumb[data-id="' + id + '"]').forEach(el => {{
        el.classList.toggle('is-pick', on);
    }});
    const mpchk = document.getElementById('modal-pick-chk');
    if (mpchk && mpchk.dataset.id === id) mpchk.checked = on;
    persistPicks();
    const pickP = document.getElementById('panel-pick');
    if (pickP && pickP.classList.contains('active')) {{
        renderPickPanel();
    }}
    if (sb && currentUser) {{
        try {{
            if (on) {{
                const {{ error }} = await sb.from('designer_picks').upsert(
                    {{ user_email: currentUser.email, item_id: id }},
                    {{ onConflict: 'user_email,item_id' }}
                );
                if (error) throw error;
                if (!remoteDesignerPicks.has(id)) remoteDesignerPicks.set(id, new Set());
                remoteDesignerPicks.get(id).add(currentUser.email);
            }} else {{
                const {{ error }} = await sb.from('designer_picks').delete()
                    .eq('user_email', currentUser.email).eq('item_id', id);
                if (error) throw error;
                if (remoteDesignerPicks.has(id)) {{
                    remoteDesignerPicks.get(id).delete(currentUser.email);
                    if (!remoteDesignerPicks.get(id).size) remoteDesignerPicks.delete(id);
                }}
            }}
            renderAttributionBadgesFor(id);
        }} catch (e) {{
            console.error('[designer_picks sync error]', e);
            showToast('동기화 실패 — 네트워크 확인 필요', true);
        }}
    }}
}}

function restoreAllPickMarks() {{
    document.querySelectorAll('.thumb').forEach(el => {{
        const on = designerPicks.has(el.dataset.id);
        el.classList.toggle('is-pick', on);
    }});
    updatePickCount();
}}

// 체크박스 이벤트 (위임)
document.addEventListener('change', (e) => {{
    const el = e.target;
    if (el.classList && el.classList.contains('item-chk')) {{
        applySelected(el.dataset.id, el.checked);
    }} else if (el.id === 'modal-chk') {{
        applySelected(el.dataset.id, el.checked);
    }} else if (el.id === 'modal-pick-chk') {{
        applyPick(el.dataset.id, el.checked);
    }}
}});

// ── Tabs ──
document.querySelectorAll('.tab-btn').forEach(btn => {{
    btn.addEventListener('click', () => {{
        const tab = btn.dataset.tab;
        document.querySelectorAll('.tab-btn').forEach(b => b.classList.toggle('active', b === btn));
        document.querySelectorAll('.panel').forEach(p =>
            p.classList.toggle('active', p.dataset.panel === tab));
        const selP = document.getElementById('panel-selected');
        const pickP = document.getElementById('panel-pick');
        selP.classList.toggle('active', tab === '__selected__');
        pickP.classList.toggle('active', tab === '__pick__');
        if (tab === '__selected__') renderSelectedPanel();
        else if (tab === '__pick__') renderPickPanel();
    }});
}});

// ── Modal ──
const modal = document.getElementById('modal');
const modalImg = document.getElementById('modal-img');
const modalTitle = document.getElementById('modal-title');
const modalKicker = document.getElementById('modal-kicker');
const modalTags = document.getElementById('modal-tags');
const modalChk = document.getElementById('modal-chk');
const modalLink = document.getElementById('modal-link');
const modalLoading = document.getElementById('modal-loading');

function openModal(id) {{
    const it = ITEMS[id];
    if (!it) return;
    modalKicker.textContent = it.tab_label + ' · ' + it.line_label + ' · ' + it.fit_label;
    modalTitle.textContent = it.name || '—';
    modalTags.innerHTML = '';
    const tagBrand = document.createElement('span');
    tagBrand.className = 'modal-tag brand';
    tagBrand.textContent = it.brand_label;
    modalTags.appendChild(tagBrand);
    if (it.sub) {{
        const s = document.createElement('span');
        s.className = 'modal-tag'; s.textContent = it.sub;
        modalTags.appendChild(s);
    }}
    if (it.match === 'core') {{
        const s = document.createElement('span');
        s.className = 'modal-tag star'; s.textContent = '⭐ ST 직접 부합';
        modalTags.appendChild(s);
    }} else if (it.match === 'adapt') {{
        const s = document.createElement('span');
        s.className = 'modal-tag adapt'; s.textContent = '✨ 타키니화 가능';
        modalTags.appendChild(s);
    }}
    modalChk.dataset.id = id;
    modalChk.checked = selected.has(id);
    const modalPickChk = document.getElementById('modal-pick-chk');
    modalPickChk.dataset.id = id;
    modalPickChk.checked = designerPicks.has(id);
    if (it.product_url) {{
        modalLink.href = it.product_url;
        modalLink.style.display = '';
    }} else {{
        modalLink.style.display = 'none';
    }}
    // 원본(hires) 이미지 로드 — 실패 시 썸네일 폴백
    modalLoading.style.display = '';
    modalImg.style.display = 'none';
    modalImg.src = '';
    const hires = it.image_hires || it.image_thumb;
    const tmpImg = new Image();
    tmpImg.onload = () => {{
        modalImg.src = hires;
        modalImg.style.display = '';
        modalLoading.style.display = 'none';
    }};
    tmpImg.onerror = () => {{
        // hires 실패 시 썸네일로 폴백
        modalImg.src = it.image_thumb;
        modalImg.style.display = '';
        modalLoading.style.display = 'none';
    }};
    tmpImg.src = hires;
    modal.classList.add('open');
    modal.setAttribute('aria-hidden', 'false');
}}

function closeModal() {{
    modal.classList.remove('open');
    modal.setAttribute('aria-hidden', 'true');
    modalImg.src = '';
}}

document.getElementById('modal-close').addEventListener('click', closeModal);
modal.addEventListener('click', (e) => {{ if (e.target === modal) closeModal(); }});
document.addEventListener('keydown', (e) => {{
    if (e.key === 'Escape' && modal.classList.contains('open')) closeModal();
}});

// 썸네일 클릭 → 모달
document.addEventListener('click', (e) => {{
    const thumb = e.target.closest('.thumb');
    if (thumb && !e.target.closest('.chk')) {{
        e.preventDefault();
        openModal(thumb.dataset.id);
    }}
}});
// 키보드 접근성 (Enter/Space로 열기)
document.addEventListener('keydown', (e) => {{
    if ((e.key === 'Enter' || e.key === ' ') && e.target.classList && e.target.classList.contains('thumb')) {{
        e.preventDefault();
        openModal(e.target.dataset.id);
    }}
}});

// ── 선택된 항목 탭 렌더 ──
function renderSelectedPanel() {{
    const body = document.getElementById('selected-body');
    if (!selected.size) {{
        body.innerHTML = '<div class="sel-empty">아직 MD PICK 된 항목이 없습니다. 각 탭에서 체크박스를 클릭하거나 모달 안에서 MD PICK 하세요.</div>';
        return;
    }}
    // 탭별로 그룹화
    const groups = {{}};
    const tabOrder = {tab_order_json};
    [...selected].forEach(id => {{
        const it = ITEMS[id];
        if (!it) return;
        if (!groups[it.tab]) groups[it.tab] = {{ label: it.tab_label, items: [] }};
        groups[it.tab].items.push(it);
    }});
    const parts = [];
    tabOrder.forEach(tk => {{
        if (!groups[tk]) return;
        const g = groups[tk];
        const thumbs = g.items.map(it => {{
            let badge = '';
            if (it.match === 'core') badge = '<span class="badge badge-core">⭐</span>';
            else if (it.match === 'adapt') badge = '<span class="badge badge-adapt">✨</span>';
            return '<div class="thumb is-selected" data-id="' + it.id + '" role="button" tabindex="0">' +
                '<img src="' + it.image_thumb + '" alt="" loading="lazy">' + badge +
                '<label class="chk" onclick="event.stopPropagation()">' +
                '<input type="checkbox" class="item-chk" data-id="' + it.id + '" checked></label>' +
                '</div>';
        }}).join('');
        parts.push('<div class="sel-group">' +
            '<div class="sel-group-head">' +
            '<span class="sel-group-title">' + g.label + '</span>' +
            '<span class="sel-group-count">' + g.items.length + ' ITEMS</span>' +
            '</div>' +
            '<div class="sel-gallery">' + thumbs + '</div>' +
            '</div>');
    }});
    body.innerHTML = parts.join('');
}}

// ── 디자이너 PICK 탭 렌더 (탭 → 브랜드 2단 그룹화) ──
function renderPickPanel() {{
    const body = document.getElementById('pick-body');
    if (!designerPicks.size) {{
        body.innerHTML = '<div class="sel-empty">아직 디자이너 PICK으로 지정된 항목이 없습니다. 썸네일 클릭 → 모달의 "🎨 디자이너 PICK" 체크박스로 지정하세요.</div>';
        return;
    }}
    const groups = {{}};
    const tabOrder = {tab_order_json};
    [...designerPicks].forEach(id => {{
        const it = ITEMS[id];
        if (!it) return;
        if (!groups[it.tab]) groups[it.tab] = {{ label: it.tab_label, brands: {{}} }};
        const b = groups[it.tab].brands;
        if (!b[it.brand]) b[it.brand] = {{ label: it.brand_label, items: [] }};
        b[it.brand].items.push(it);
    }});
    const parts = [];
    tabOrder.forEach(tk => {{
        if (!groups[tk]) return;
        const g = groups[tk];
        Object.keys(g.brands).forEach(bk => {{
            const bg = g.brands[bk];
            const thumbs = bg.items.map(it => {{
                let badge = '';
                if (it.match === 'core') badge = '<span class="badge badge-core">⭐</span>';
                else if (it.match === 'adapt') badge = '<span class="badge badge-adapt">✨</span>';
                const selCls = selected.has(it.id) ? ' is-selected' : '';
                return '<div class="thumb is-pick' + selCls + '" data-id="' + it.id + '" role="button" tabindex="0">' +
                    '<img src="' + it.image_thumb + '" alt="" loading="lazy">' + badge + '</div>';
            }}).join('');
            parts.push('<div class="sel-group">' +
                '<div class="sel-group-head">' +
                '<span class="sel-group-title">' + g.label + ' · ' + bg.label + '</span>' +
                '<span class="sel-group-count">' + bg.items.length + ' ITEMS</span>' +
                '</div>' +
                '<div class="sel-gallery">' + thumbs + '</div>' +
                '</div>');
        }});
    }});
    body.innerHTML = parts.join('');
}}

// ── 디자이너 PICK 백업 JSON ──
function downloadPickBackup() {{
    const arr = [...designerPicks].map(id => ITEMS[id]).filter(Boolean).map(it => ({{
        id: it.id, brand: it.brand_label, brand_key: it.brand,
        line: it.line_label, tab: it.tab_label,
        fit: it.fit_label, sub: it.sub,
        name: it.name, product_url: it.product_url,
        image_hires: it.image_hires,
        image_thumb: it.image_thumb,
        st_mark: it.match === 'core' ? '⭐ 직접 부합'
               : it.match === 'adapt' ? '✨ 타키니화 가능' : '',
    }}));
    const payload = {{
        generated_at: new Date().toISOString(),
        kind: 'designer_pick',
        count: arr.length, total: TOTAL_ITEMS,
        items: arr,
    }};
    const blob = new Blob([JSON.stringify(payload, null, 2)], {{type: 'application/json'}});
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    const ts = new Date().toISOString().slice(0,19).replace(/[:T]/g,'-');
    a.href = url; a.download = 'designer_pick_backup_' + ts + '.json';
    a.click();
    URL.revokeObjectURL(url);
    showToast('PICK 백업 JSON — ' + arr.length + '건');
}}

async function downloadPickZip() {{
    if (!designerPicks.size) {{ showToast('PICK 항목이 없습니다', true); return; }}
    await downloadIdsAsZip(designerPicks, document.getElementById('btn-export-pick-zip'), 'designer_pick');
}}

async function clearAllPicks() {{
    if (!designerPicks.size) {{ showToast('PICK 항목이 없습니다'); return; }}
    if (SUPABASE_ENABLED && !currentUser) {{ showToast('로그인 후 사용할 수 있습니다', true); return; }}
    if (!confirm('디자이너 PICK ' + designerPicks.size + '건을 모두 해제하시겠습니까?')) return;
    const ids = [...designerPicks];
    designerPicks.clear();
    persistPicks();
    ids.forEach(id => {{
        document.querySelectorAll('.thumb[data-id="' + id + '"]').forEach(el => el.classList.remove('is-pick'));
    }});
    if (document.getElementById('panel-pick').classList.contains('active')) {{
        renderPickPanel();
    }}
    if (sb && currentUser) {{
        try {{
            const {{ error }} = await sb.from('designer_picks').delete().eq('user_email', currentUser.email);
            if (error) throw error;
            ids.forEach(id => {{
                if (remoteDesignerPicks.has(id)) {{
                    remoteDesignerPicks.get(id).delete(currentUser.email);
                    if (!remoteDesignerPicks.get(id).size) remoteDesignerPicks.delete(id);
                }}
                renderAttributionBadgesFor(id);
            }});
        }} catch (e) {{
            console.error('[clear designer_picks error]', e);
            showToast('동기화 실패 — 새로고침하여 확인하세요', true);
        }}
    }}
    showToast('PICK 전체 해제 완료');
}}

// ── 백업 JSON 다운로드 ──
function downloadBackup() {{
    const arr = [...selected].map(id => ITEMS[id]).filter(Boolean).map(it => ({{
        id: it.id, brand: it.brand_label, brand_key: it.brand,
        line: it.line_label, tab: it.tab_label,
        fit: it.fit_label, sub: it.sub,
        name: it.name, product_url: it.product_url,
        image_hires: it.image_hires,  // 원본 URL / file:// 경로 (Python 스크립트용)
        image_thumb: it.image_thumb,  // 폴백용 data URL
        st_mark: it.match === 'core' ? '⭐ 직접 부합'
               : it.match === 'adapt' ? '✨ 타키니화 가능' : '',
    }}));
    const payload = {{
        generated_at: new Date().toISOString(),
        count: arr.length, total: TOTAL_ITEMS,
        items: arr,
    }};
    const blob = new Blob([JSON.stringify(payload, null, 2)], {{type: 'application/json'}});
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    const ts = new Date().toISOString().slice(0,19).replace(/[:T]/g,'-');
    a.href = url; a.download = 'md_pick_backup_' + ts + '.json';
    a.click();
    URL.revokeObjectURL(url);
    showToast('MD PICK 백업 JSON 다운로드 — ' + arr.length + '건');
}}

// ── PNG ZIP 다운로드 ──
//   ids: 다운로드할 item id 집합(Set 또는 Array). btnEl: 상태 표시용 버튼 엘리먼트. prefix: 파일명 접두사
async function downloadIdsAsZip(ids, btnEl, prefix) {{
    const idArr = [...ids];
    if (!idArr.length) {{ showToast('다운로드할 항목이 없습니다', true); return; }}
    if (typeof JSZip === 'undefined') {{
        showToast('JSZip 라이브러리 로드 실패 — 네트워크 확인 필요', true); return;
    }}
    const orig = btnEl.textContent;
    btnEl.disabled = true;
    btnEl.textContent = 'ZIP 생성 중 (0/' + idArr.length + ')...';
    const zip = new JSZip();
    const rows = [['id','브랜드','라인','탭','핏','소분류','상품명','상품URL','이미지파일','ST마크']];
    let i = 0;
    for (const id of idArr) {{
        const it = ITEMS[id]; if (!it) continue;
        i++;
        btnEl.textContent = 'ZIP 생성 중 (' + i + '/' + idArr.length + ')...';
        let imgBlob = null;
        const urls = [it.image_hires, it.image_thumb].filter(Boolean);
        for (const u of urls) {{
            try {{
                const resp = await fetch(u);
                if (resp.ok) {{ imgBlob = await resp.blob(); break; }}
            }} catch (e) {{ /* 다음 fallback */ }}
        }}
        if (!imgBlob) continue;
        const safeName = (it.brand_label + '_' + it.name)
            .replace(/[\\\\/:*?"<>|]/g, '_').slice(0, 80);
        const ext = imgBlob.type.includes('png') ? 'png' : 'jpg';
        const fname = String(i).padStart(4,'0') + '_' + it.tab_label + '_' + safeName + '.' + ext;
        zip.file(fname, imgBlob);
        const mark = it.match === 'core' ? '⭐ 직접 부합'
                   : it.match === 'adapt' ? '✨ 타키니화 가능' : '';
        rows.push([it.id, it.brand_label, it.line_label, it.tab_label, it.fit_label,
                   it.sub, it.name, it.product_url, fname, mark]);
    }}
    const csv = '\\ufeff' + rows.map(r => r.map(c =>
        '"' + String(c ?? '').replace(/"/g,'""') + '"').join(',')).join('\\r\\n');
    zip.file('manifest.csv', csv);
    const content = await zip.generateAsync({{type: 'blob', compression: 'DEFLATE'}});
    const url = URL.createObjectURL(content);
    const a = document.createElement('a');
    const ts = new Date().toISOString().slice(0,19).replace(/[:T]/g,'-');
    a.href = url; a.download = prefix + '_' + ts + '.zip';
    a.click();
    URL.revokeObjectURL(url);
    btnEl.disabled = false; btnEl.textContent = orig;
    showToast('이미지 ZIP 다운로드 완료 — ' + idArr.length + '건');
}}

async function downloadImagesZip() {{
    if (!selected.size) {{ showToast('MD PICK 된 항목이 없습니다', true); return; }}
    await downloadIdsAsZip(selected, document.getElementById('btn-export-xlsx'), 'md_pick_images');
}}

// ⭐✨ 추천 항목 전체 다운로드 (선택 불필요)
async function downloadRecommendedZip() {{
    const recIds = Object.keys(ITEMS).filter(id =>
        ITEMS[id].match === 'core' || ITEMS[id].match === 'adapt');
    if (!recIds.length) {{ showToast('추천 항목이 없습니다', true); return; }}
    if (!confirm('⭐✨ 추천 ' + recIds.length + '건 전체 이미지를 다운로드합니다. 계속하시겠습니까?')) return;
    await downloadIdsAsZip(recIds, document.getElementById('btn-download-rec'), 'recommended_all');
}}

// 🔥 고화질용 백업 — ⭐✨ 추천 항목 + (선택된 경우) 체크박스 항목 합집합을 JSON으로 추출
//    Python download_hires.py 로 원본 해상도 다운로드용
function downloadHiresBackup() {{
    const recIds = new Set(Object.keys(ITEMS).filter(id =>
        ITEMS[id].match === 'core' || ITEMS[id].match === 'adapt'));
    // 사용자 선택도 합집합 (추가 반영)
    selected.forEach(id => recIds.add(id));
    const arr = [...recIds].map(id => ITEMS[id]).filter(Boolean).map(it => ({{
        id: it.id, brand: it.brand_label, brand_key: it.brand,
        line: it.line_label, tab: it.tab_label, fit: it.fit_label,
        sub: it.sub, name: it.name, product_url: it.product_url,
        image_hires: it.image_hires, image_thumb: it.image_thumb,
        st_mark: it.match === 'core' ? '⭐ 직접 부합'
               : it.match === 'adapt' ? '✨ 타키니화 가능' : '(선택)',
    }}));
    const payload = {{
        generated_at: new Date().toISOString(),
        count: arr.length,
        recommended_count: [...recIds].filter(id => ITEMS[id] && (ITEMS[id].match === 'core' || ITEMS[id].match === 'adapt')).length,
        selected_count: selected.size,
        items: arr,
        usage: "Run: python download_hires.py <이_파일명> (원본 해상도 이미지 ZIP 생성)",
    }};
    const blob = new Blob([JSON.stringify(payload, null, 2)], {{type: 'application/json'}});
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    const ts = new Date().toISOString().slice(0,19).replace(/[:T]/g,'-');
    a.href = url; a.download = 'hires_backup_' + ts + '.json';
    a.click();
    URL.revokeObjectURL(url);
    showToast('백업 완료 ' + arr.length + '건 — 터미널에서: python download_hires.py hires_backup_' + ts + '.json');
}}

// ── 독립 다중 필터: 레전드 3박스 (core · adapt · exc) 각각 토글 — 복수 활성 가능 ──
const FILTER_MODES_KEY = 'line_matrix_filter_modes_v1';
let filterModes = new Set();
try {{
    const raw = localStorage.getItem(FILTER_MODES_KEY);
    if (raw) filterModes = new Set(JSON.parse(raw).filter(m => ['core','adapt','exc'].includes(m)));
}} catch (e) {{ console.warn('filter modes read failed:', e); }}
// 구버전 localStorage 키 정리
['line_matrix_filter_rec_v1','line_matrix_filter_exc_v1','line_matrix_filter_mode_v1']
    .forEach(k => localStorage.removeItem(k));

function persistFilterModes() {{
    if (filterModes.size) localStorage.setItem(FILTER_MODES_KEY, JSON.stringify([...filterModes]));
    else localStorage.removeItem(FILTER_MODES_KEY);
}}

function applyFilterState() {{
    document.body.classList.toggle('has-filter', filterModes.size > 0);
    ['core','adapt','exc'].forEach(m => {{
        document.body.classList.toggle('show-' + m, filterModes.has(m));
    }});
    document.querySelectorAll('.legend-item').forEach(el => {{
        el.classList.toggle('active', filterModes.has(el.dataset.filter));
    }});
    updateCellCounts();
}}

function toggleFilterMode(mode) {{
    if (filterModes.has(mode)) filterModes.delete(mode);
    else filterModes.add(mode);
    persistFilterModes();
    applyFilterState();
}}

// 필터 상태에 따라 각 cell의 count pill 업데이트
function updateCellCounts() {{
    const anyFilter = filterModes.size > 0;
    document.querySelectorAll('.gallery-cell').forEach(cell => {{
        const thumbs = cell.querySelectorAll('.thumb');
        const total = thumbs.length;
        let visible = total;
        if (anyFilter) {{
            visible = Array.from(thumbs).filter(t => {{
                const m = t.dataset.match;
                if (filterModes.has('core')  && m === 'core')  return true;
                if (filterModes.has('adapt') && m === 'adapt') return true;
                if (filterModes.has('exc')   && m === '')      return true;
                return false;
            }}).length;
        }}
        const pill = cell.querySelector('.count-pill');
        if (pill) {{
            pill.textContent = anyFilter && total > 0 ? (visible + '/' + total) : total;
        }}
        cell.classList.toggle('hidden-all', anyFilter && visible === 0 && total > 0);
    }});
}}

// 레전드 박스 클릭/키보드 핸들러
document.querySelectorAll('.legend-item').forEach(el => {{
    el.addEventListener('click', () => toggleFilterMode(el.dataset.filter));
    el.addEventListener('keydown', (e) => {{
        if (e.key === 'Enter' || e.key === ' ') {{
            e.preventDefault();
            toggleFilterMode(el.dataset.filter);
        }}
    }});
}});

// ── 전체 해제 ──
async function clearAllSelections() {{
    if (!selected.size) {{ showToast('MD PICK 된 항목이 없습니다'); return; }}
    if (SUPABASE_ENABLED && !currentUser) {{ showToast('로그인 후 사용할 수 있습니다', true); return; }}
    if (!confirm('MD PICK ' + selected.size + '건을 모두 해제하시겠습니까?')) return;
    const ids = [...selected];
    selected.clear();
    persist();
    ids.forEach(id => {{
        document.querySelectorAll('input.item-chk[data-id="' + id + '"]').forEach(el => el.checked = false);
        document.querySelectorAll('.thumb[data-id="' + id + '"]').forEach(el => el.classList.remove('is-selected'));
    }});
    if (document.getElementById('panel-selected').classList.contains('active')) {{
        renderSelectedPanel();
    }}
    if (sb && currentUser) {{
        try {{
            const {{ error }} = await sb.from('md_picks').delete().eq('user_email', currentUser.email);
            if (error) throw error;
            ids.forEach(id => {{
                if (remoteMdPicks.has(id)) {{
                    remoteMdPicks.get(id).delete(currentUser.email);
                    if (!remoteMdPicks.get(id).size) remoteMdPicks.delete(id);
                }}
                renderAttributionBadgesFor(id);
            }});
        }} catch (e) {{
            console.error('[clear md_picks error]', e);
            showToast('동기화 실패 — 새로고침하여 확인하세요', true);
        }}
    }}
    showToast('MD PICK 전체 해제 완료');
}}

document.getElementById('btn-backup').addEventListener('click', downloadBackup);
document.getElementById('btn-export-xlsx').addEventListener('click', downloadImagesZip);
document.getElementById('btn-clear-all').addEventListener('click', clearAllSelections);
document.getElementById('btn-download-rec').addEventListener('click', downloadRecommendedZip);
document.getElementById('btn-backup-hires').addEventListener('click', downloadHiresBackup);
document.getElementById('btn-backup-pick').addEventListener('click', downloadPickBackup);
document.getElementById('btn-export-pick-zip').addEventListener('click', downloadPickZip);
document.getElementById('btn-clear-pick').addEventListener('click', clearAllPicks);

// ── 이미지 확대/축소 (썸네일 배율) ──
// --thumb-w CSS 변수로 갤러리 영역의 .thumb 크기만 조절 (레이아웃 전체는 불변)
const ZOOM_KEY = 'line_matrix_zoom_v1';
const zoomSlider = document.getElementById('zoom-slider');
const zoomVal = document.getElementById('zoom-val');
const ZOOM_MIN = 60, ZOOM_MAX = 280, ZOOM_STEP = 10;

function applyZoom(px) {{
    px = Math.max(ZOOM_MIN, Math.min(ZOOM_MAX, px));
    document.documentElement.style.setProperty('--thumb-w', px + 'px');
    zoomSlider.value = px;
    zoomVal.textContent = Math.round(px / 100 * 100) + '%';  // 100px = 100%
    localStorage.setItem(ZOOM_KEY, String(px));
}}

zoomSlider.addEventListener('input', () => applyZoom(parseInt(zoomSlider.value, 10)));
document.getElementById('zoom-in').addEventListener('click',
    () => applyZoom(parseInt(zoomSlider.value, 10) + ZOOM_STEP));
document.getElementById('zoom-out').addEventListener('click',
    () => applyZoom(parseInt(zoomSlider.value, 10) - ZOOM_STEP));

// 초기값 복원
const savedZoom = parseInt(localStorage.getItem(ZOOM_KEY) || '100', 10);
applyZoom(savedZoom);

// 초기화
restoreAllCheckboxes();
restoreAllPickMarks();
applyFilterState();

// ── Supabase Auth · Realtime · Attribution ────────────────────────────
async function initSupabase() {{
    if (!SUPABASE_ENABLED) {{
        console.info('[Supabase] 설정 없음 — localStorage 전용 모드');
        return;
    }}
    if (!sb) {{
        console.warn('[Supabase] SDK 로드 실패');
        return;
    }}
    const {{ data: {{ session }} }} = await sb.auth.getSession();
    if (session && session.user) {{
        await onLogin(session.user);
    }} else {{
        showAuthModal();
    }}
    sb.auth.onAuthStateChange((event, s) => {{
        if (event === 'SIGNED_IN' && s && s.user) onLogin(s.user);
        else if (event === 'SIGNED_OUT') onLogout();
    }});
}}

async function onLogin(user) {{
    currentUser = {{ email: user.email }};
    hideAuthModal();
    showUserPill(user.email);
    await loadRemoteAll();
    subscribeRealtime();
}}

function onLogout() {{
    currentUser = null;
    hideUserPill();
    selected.clear(); designerPicks.clear();
    remoteMdPicks.clear(); remoteDesignerPicks.clear();
    restoreAllCheckboxes();
    restoreAllPickMarks();
    document.querySelectorAll('.attr-badge').forEach(el => el.remove());
    showAuthModal();
}}

async function loadRemoteAll() {{
    try {{
        const {{ data: md, error: e1 }} = await sb.from('md_picks').select('user_email,item_id');
        if (e1) throw e1;
        remoteMdPicks.clear(); selected.clear();
        (md || []).forEach(r => {{
            if (!remoteMdPicks.has(r.item_id)) remoteMdPicks.set(r.item_id, new Set());
            remoteMdPicks.get(r.item_id).add(r.user_email);
            if (currentUser && r.user_email === currentUser.email) selected.add(r.item_id);
        }});
        const {{ data: dp, error: e2 }} = await sb.from('designer_picks').select('user_email,item_id');
        if (e2) throw e2;
        remoteDesignerPicks.clear(); designerPicks.clear();
        (dp || []).forEach(r => {{
            if (!remoteDesignerPicks.has(r.item_id)) remoteDesignerPicks.set(r.item_id, new Set());
            remoteDesignerPicks.get(r.item_id).add(r.user_email);
            if (currentUser && r.user_email === currentUser.email) designerPicks.add(r.item_id);
        }});
        restoreAllCheckboxes();
        restoreAllPickMarks();
        renderAllAttributionBadges();
        updateSelCount(); updatePickCount();
    }} catch (e) {{
        console.error('[loadRemoteAll]', e);
        showToast('데이터 로드 실패: ' + (e.message || e), true);
    }}
}}

function subscribeRealtime() {{
    if (!sb) return;
    sb.channel('md_picks_rt')
      .on('postgres_changes', {{event: '*', schema: 'public', table: 'md_picks'}}, handleMdPickChange)
      .subscribe();
    sb.channel('designer_picks_rt')
      .on('postgres_changes', {{event: '*', schema: 'public', table: 'designer_picks'}}, handleDesignerPickChange)
      .subscribe();
}}

function handleMdPickChange(payload) {{
    const row = payload.new && Object.keys(payload.new).length ? payload.new : payload.old;
    if (!row || !row.item_id) return;
    const itemId = row.item_id, email = row.user_email;
    const isMine = currentUser && email === currentUser.email;
    if (payload.eventType === 'INSERT') {{
        if (!remoteMdPicks.has(itemId)) remoteMdPicks.set(itemId, new Set());
        remoteMdPicks.get(itemId).add(email);
        if (isMine) {{ selected.add(itemId); syncThumbSelected(itemId, true); }}
    }} else if (payload.eventType === 'DELETE') {{
        if (remoteMdPicks.has(itemId)) {{
            remoteMdPicks.get(itemId).delete(email);
            if (!remoteMdPicks.get(itemId).size) remoteMdPicks.delete(itemId);
        }}
        if (isMine) {{ selected.delete(itemId); syncThumbSelected(itemId, false); }}
    }}
    renderAttributionBadgesFor(itemId);
    updateSelCount();
    if (isMine && document.getElementById('panel-selected').classList.contains('active')) {{
        renderSelectedPanel();
    }}
}}

function handleDesignerPickChange(payload) {{
    const row = payload.new && Object.keys(payload.new).length ? payload.new : payload.old;
    if (!row || !row.item_id) return;
    const itemId = row.item_id, email = row.user_email;
    const isMine = currentUser && email === currentUser.email;
    if (payload.eventType === 'INSERT') {{
        if (!remoteDesignerPicks.has(itemId)) remoteDesignerPicks.set(itemId, new Set());
        remoteDesignerPicks.get(itemId).add(email);
        if (isMine) {{ designerPicks.add(itemId); syncThumbPick(itemId, true); }}
    }} else if (payload.eventType === 'DELETE') {{
        if (remoteDesignerPicks.has(itemId)) {{
            remoteDesignerPicks.get(itemId).delete(email);
            if (!remoteDesignerPicks.get(itemId).size) remoteDesignerPicks.delete(itemId);
        }}
        if (isMine) {{ designerPicks.delete(itemId); syncThumbPick(itemId, false); }}
    }}
    renderAttributionBadgesFor(itemId);
    updatePickCount();
    if (isMine && document.getElementById('panel-pick').classList.contains('active')) {{
        renderPickPanel();
    }}
}}

function syncThumbSelected(itemId, on) {{
    document.querySelectorAll('.thumb[data-id="' + itemId + '"]').forEach(el => el.classList.toggle('is-selected', on));
    document.querySelectorAll('input.item-chk[data-id="' + itemId + '"]').forEach(el => el.checked = on);
    const mchk = document.getElementById('modal-chk');
    if (mchk && mchk.dataset.id === itemId) mchk.checked = on;
}}

function syncThumbPick(itemId, on) {{
    document.querySelectorAll('.thumb[data-id="' + itemId + '"]').forEach(el => el.classList.toggle('is-pick', on));
    const mpchk = document.getElementById('modal-pick-chk');
    if (mpchk && mpchk.dataset.id === itemId) mpchk.checked = on;
}}

// Attribution 배지: 썸네일 좌하단 "👤N" + tooltip (누가 MD/PICK 했는지)
function renderAttributionBadgesFor(itemId) {{
    const mdSet = remoteMdPicks.get(itemId) || new Set();
    const pickSet = remoteDesignerPicks.get(itemId) || new Set();
    const allEmails = new Set([...mdSet, ...pickSet]);
    document.querySelectorAll('.thumb[data-id="' + itemId + '"]').forEach(el => {{
        let attr = el.querySelector(':scope > .attr-badge');
        if (!allEmails.size) {{ if (attr) attr.remove(); return; }}
        if (!attr) {{
            attr = document.createElement('div');
            attr.className = 'attr-badge';
            el.appendChild(attr);
        }}
        const hasMe = currentUser && allEmails.has(currentUser.email);
        const inner = (hasMe ? '<span class="me-dot"></span>' : '') + '👤' + allEmails.size;
        attr.innerHTML = inner;
        const lines = [...allEmails].map(e => {{
            const tags = [];
            if (mdSet.has(e)) tags.push('MD');
            if (pickSet.has(e)) tags.push('PICK');
            return e.split('@')[0] + ' (' + tags.join('+') + ')';
        }});
        attr.title = lines.join('\\n');
    }});
}}

function renderAllAttributionBadges() {{
    document.querySelectorAll('.attr-badge').forEach(el => el.remove());
    const ids = new Set([...remoteMdPicks.keys(), ...remoteDesignerPicks.keys()]);
    ids.forEach(renderAttributionBadgesFor);
}}

// Auth UI helpers
function showAuthModal() {{
    const m = document.getElementById('auth-modal');
    m.classList.add('open');
    m.setAttribute('aria-hidden', 'false');
    setTimeout(() => {{ const i = document.getElementById('auth-email'); if (i) i.focus(); }}, 100);
}}
function hideAuthModal() {{
    const m = document.getElementById('auth-modal');
    m.classList.remove('open');
    m.setAttribute('aria-hidden', 'true');
}}
function showUserPill(email) {{
    const pill = document.getElementById('user-pill');
    document.getElementById('user-pill-email').textContent = email;
    pill.style.display = '';
}}
function hideUserPill() {{ document.getElementById('user-pill').style.display = 'none'; }}

async function sendMagicLink() {{
    const emailEl = document.getElementById('auth-email');
    const statusEl = document.getElementById('auth-status');
    const btn = document.getElementById('auth-send');
    const email = (emailEl.value || '').trim().toLowerCase();
    if (!email) {{
        statusEl.textContent = '이메일을 입력해주세요'; statusEl.className = 'auth-status error'; return;
    }}
    if (!email.endsWith(SUPABASE_CFG.domain)) {{
        statusEl.textContent = SUPABASE_CFG.domain + ' 이메일만 허용됩니다';
        statusEl.className = 'auth-status error'; return;
    }}
    btn.disabled = true;
    statusEl.textContent = '전송 중...'; statusEl.className = 'auth-status';
    const redirect = window.location.origin + window.location.pathname;
    const {{ error }} = await sb.auth.signInWithOtp({{
        email,
        options: {{ emailRedirectTo: redirect }}
    }});
    btn.disabled = false;
    if (error) {{
        statusEl.textContent = '전송 실패: ' + error.message;
        statusEl.className = 'auth-status error';
    }} else {{
        statusEl.innerHTML = '✉️ <strong>' + email + '</strong> 로 링크 전송 — 메일함을 확인하세요';
        statusEl.className = 'auth-status ok';
    }}
}}

async function doLogout() {{
    if (!sb) return;
    await sb.auth.signOut();
}}

// Auth UI 이벤트 바인딩
(function bindAuthEvents() {{
    const sendBtn = document.getElementById('auth-send');
    if (sendBtn) sendBtn.addEventListener('click', sendMagicLink);
    const emailInp = document.getElementById('auth-email');
    if (emailInp) emailInp.addEventListener('keydown', e => {{
        if (e.key === 'Enter') {{ e.preventDefault(); sendMagicLink(); }}
    }});
    const skipBtn = document.getElementById('auth-skip');
    if (skipBtn) skipBtn.addEventListener('click', () => {{
        hideAuthModal();
        showToast('읽기 전용 모드 — 변경사항은 저장되지 않습니다', true);
    }});
    const logoutBtn = document.getElementById('user-pill-logout');
    if (logoutBtn) logoutBtn.addEventListener('click', doLogout);
}})();

// Supabase 부팅
initSupabase();

// ── Toast ──
const toastEl = document.getElementById('toast');
let toastTimer;
function showToast(msg, isError = false) {{
    clearTimeout(toastTimer);
    toastEl.textContent = msg;
    toastEl.classList.toggle('error', isError);
    toastEl.classList.add('show');
    toastTimer = setTimeout(() => toastEl.classList.remove('show'), 2500);
}}

// ── Capture (clipboard / download) ──
async function capturePanel(panel, mode, btn) {{
    const scroll = panel.querySelector('.matrix-scroll');
    const head = panel.querySelector('.panel-head');

    // Full-render target wrapper (temporary)
    const wrap = document.createElement('div');
    wrap.style.cssText = 'display:inline-block;background:#ffffff;padding:0;';
    const headClone = head.cloneNode(true);
    headClone.style.cssText = 'padding:18px 22px;border:1px solid #e5e7ec;border-radius:14px 14px 0 0;border-bottom:0;background:#fff;';
    // hide action buttons in clone
    const actClone = headClone.querySelector('.panel-actions');
    if (actClone) actClone.remove();
    wrap.appendChild(headClone);

    const matrixClone = scroll.cloneNode(true);
    matrixClone.style.cssText = 'overflow:visible;border:1px solid #e5e7ec;border-top:0;border-radius:0 0 14px 14px;background:#fff;';
    // 캡처 시 체크박스 · 선택/PICK 하이라이트 제거 (상품 이미지만 깔끔하게)
    matrixClone.querySelectorAll('.chk').forEach(el => el.remove());
    matrixClone.querySelectorAll('.thumb.is-selected').forEach(el => el.classList.remove('is-selected'));
    matrixClone.querySelectorAll('.thumb.is-pick').forEach(el => el.classList.remove('is-pick'));
    wrap.appendChild(matrixClone);

    document.body.appendChild(wrap);
    wrap.style.position = 'fixed';
    wrap.style.left = '-99999px';
    wrap.style.top = '0';

    btn.disabled = true;
    const origText = btn.querySelector('span').textContent;
    btn.querySelector('span').textContent = '캡처 중...';

    try {{
        const canvas = await html2canvas(wrap, {{
            backgroundColor: '#ffffff',
            useCORS: true,
            scale: 2,
            logging: false,
        }});
        wrap.remove();

        if (mode === 'copy') {{
            canvas.toBlob(async (blob) => {{
                try {{
                    await navigator.clipboard.write([new ClipboardItem({{[blob.type]: blob}})]);
                    showToast('클립보드에 복사되었습니다');
                }} catch (e) {{
                    showToast('클립보드 복사 실패 — 이미지 저장으로 이용해주세요', true);
                }}
            }}, 'image/png');
        }} else {{
            const link = document.createElement('a');
            const date = new Date().toISOString().slice(0, 10);
            link.download = `matrix_${{panel.dataset.panel}}_${{date}}.png`;
            link.href = canvas.toDataURL('image/png');
            link.click();
            showToast('이미지가 저장되었습니다');
        }}
    }} catch (e) {{
        console.error(e);
        wrap.remove();
        showToast('캡처 중 오류 발생: ' + e.message, true);
    }} finally {{
        btn.disabled = false;
        btn.querySelector('span').textContent = origText;
    }}
}}

document.querySelectorAll('.panel-actions button').forEach(btn => {{
    btn.addEventListener('click', () => {{
        const panel = btn.closest('.panel');
        capturePanel(panel, btn.dataset.action, btn);
    }});
}});
</script>
</body></html>"""


def main():
    print("[load]")
    items = (
        load_alo() + load_wilson() + load_lacoste() + load_rl()
        + load_descente() + load_lululemon() + load_celine() + load_loropiana()
        + load_skims() + load_miumiu() + load_prada() + load_sportyandrich()
    )
    print(f"  의류 상품 총 {len(items)}개")

    # 모든 원격 이미지를 data URL로 변환 (캡처 CORS 이슈 근본 해결)
    cache_all_images_to_data_url(items)
    overrides = load_fit_overrides()
    if overrides:
        print(f"  핏 오버라이드 {len(overrides)}건 로드")
    matrix = build_matrix(items, overrides)

    # 분포 출력
    print("[matrix]")
    for t, lbl, _ in TABS:
        line_totals = {lk: sum(len(matrix[t][lk][fk]) for fk, _ in fit_cols_for(t)) for lk, _, _, _ in LINES}
        print(f"  [{lbl}] 클래식={line_totals['classic']} 어슬레져={line_totals['athleisure']} 스포츠={line_totals['sport']}")

    OUT.write_text(render_html(matrix), encoding="utf-8")
    print(f"[done] {OUT}  ({OUT.stat().st_size/1024:.1f}KB)")


if __name__ == "__main__":
    main()
